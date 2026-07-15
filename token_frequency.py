"""
Token frequency report for product names.

Reads an .xlsx from the input folder, tokenises every product name, counts
token frequency, and writes:
  - "Token Frequency" sheet: TOKEN, FREQUENCY (sorted descending)
  - "Examples" sheet: each token with up to 5 example product names

Edit the CONFIG block below, then run:
    pip install -r requirements.txt
    python token_frequency.py
"""

import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook


# ===== CONFIG — set these before running =====================================

INPUT_FOLDER = "input"
INPUT_FILE = "Products_639154628564714958.xlsx"   # filename inside INPUT_FOLDER
COLUMN_NAME = "Product name"                       # header of the name column
SHEET_NAME = None                                  # None = first sheet, or a sheet name string
OUTPUT_FILE = "token_frequency.xlsx"
MAX_EXAMPLES = 10

# =============================================================================


DELIMITER_PATTERN = re.compile(r"[\s/\-_]+")
PUNCTUATION_PATTERN = re.compile(r"[^\w]")


def tokenize(text: str) -> list[str]:
    """Split on whitespace and /, -, _; strip punctuation; uppercase; filter."""
    if not text:
        return []

    upper = str(text).upper()
    raw_parts = DELIMITER_PATTERN.split(upper)

    tokens = []
    for part in raw_parts:
        cleaned = PUNCTUATION_PATTERN.sub("", part)
        if len(cleaned) < 2:
            continue
        if cleaned.isdigit():
            continue
        tokens.append(cleaned)

    return tokens


script_dir = Path(__file__).resolve().parent
input_path = script_dir / INPUT_FOLDER / INPUT_FILE
output_path = script_dir / OUTPUT_FILE

if not input_path.is_file():
    raise SystemExit(f"Input file not found: {input_path}")

print(f"Reading {input_path.name} ...")

wb_in = load_workbook(input_path, read_only=True, data_only=True)
ws_in = wb_in[SHEET_NAME] if SHEET_NAME else wb_in.active

header_row = next(ws_in.iter_rows(min_row=1, max_row=1, values_only=True))
headers = [str(h).strip() if h is not None else "" for h in header_row]

try:
    col_idx = headers.index(COLUMN_NAME)
except ValueError:
    wb_in.close()
    raise SystemExit(
        f"Column {COLUMN_NAME!r} not found.\nAvailable columns: {headers}"
    )

token_counts: Counter[str] = Counter()
token_examples: dict[str, list[str]] = defaultdict(list)

row_count = 0
for row in ws_in.iter_rows(min_row=2, values_only=True):
    if col_idx >= len(row):
        continue
    name = row[col_idx]
    if name is None or str(name).strip() == "":
        continue

    name_str = str(name).strip()
    tokens = tokenize(name_str)
    if not tokens:
        continue

    row_count += 1
    token_counts.update(tokens)

    for token in set(tokens):
        examples = token_examples[token]
        if len(examples) < MAX_EXAMPLES and name_str not in examples:
            examples.append(name_str)

wb_in.close()
print(f"  Processed {row_count:,} product names")
print(f"  Unique tokens: {len(token_counts):,}")

print(f"Writing {output_path.name} ...")

wb_out = Workbook()
ws_freq = wb_out.active
ws_freq.title = "Token Frequency"
ws_freq.append(["TOKEN", "FREQUENCY"])
for token, freq in token_counts.most_common():
    ws_freq.append([token, freq])

ws_ex = wb_out.create_sheet("Examples")
example_headers = ["TOKEN", "FREQUENCY"] + [
    f"EXAMPLE_{i}" for i in range(1, MAX_EXAMPLES + 1)
]
ws_ex.append(example_headers)

for token, freq in token_counts.most_common():
    examples = token_examples[token]
    row = [token, freq, *examples]
    row.extend([None] * (MAX_EXAMPLES - len(examples)))
    ws_ex.append(row)

wb_out.save(output_path)
print(f"Done. Output: {output_path}")
