"""
Similarity reviewer — one-match duel stack.

Run:
    pip install -r requirements.txt
    python review_gui.py

Left                = duplicate
Right               = unique
Up                  = discard
Down                = unreview
Tab                 = next child (candidate in queue)
Ctrl+Tab            = previous child
Space               = next parent (next cluster)
Ctrl+Space          = previous parent (previous cluster)
O                   = open results workbook
I                   = choose product input Excel and run similarity
                      (also prompted automatically on startup)
S / R               = Reports screen (export for superiors)
T                   = toggle dark mode
Ctrl+F              = jump to a parent (cluster reference) code
G                   = show/hide the Related (semantic) suggestions panel
Related panel       = Jump to that cluster, or Pull into this cluster (human only)

Progress is saved in separate sidecar files next to the results workbook
(never modifies the original similarity_results*.xlsx):
  review_progress__<name>.json
  review_decisions__<name>.xlsx
"""

from __future__ import annotations

import json
import re
import sys
import time
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from PySide6.QtCore import QEvent, QRect, QSize, Qt, QTimer
from PySide6.QtGui import (
    QColor,
    QFont,
    QFontMetrics,
    QKeySequence,
    QMouseEvent,
    QPainter,
    QPen,
    QShortcut,
)
from PySide6.QtWidgets import (
    QApplication,
    QDialog,
    QFrame,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QSizePolicy,
    QStackedWidget,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

from text_normalize import normalize_tokens

def app_base_dir() -> Path:
    """Project / package root (parent of Review/ when frozen)."""
    if getattr(sys, "frozen", False):
        # .../SimilarityParser/Review/Similarity Review.exe → package root
        return Path(sys.executable).resolve().parent.parent
    return Path(__file__).resolve().parent


APP_DIR = app_base_dir()
DEFAULT_RESULTS = APP_DIR / "similarity_results.xlsx"
PROGRESS_FILE = APP_DIR / "review_progress.json"  # legacy fallback
INPUT_FOLDER = APP_DIR / "input"


def _boot_log(message: str) -> None:
    """Append a line so failed silent launches are diagnosable."""
    try:
        path = APP_DIR / "review_boot.log"
        with path.open("a", encoding="utf-8") as f:
            f.write(f"{datetime.now(timezone.utc).isoformat()} {message}\n")
    except OSError:
        pass


def windows_unblock(path: Path) -> None:
    """Clear Mark-of-the-Web so Windows allows launching downloaded exes/dlls."""
    if sys.platform != "win32":
        return
    try:
        from ctypes import windll

        windll.kernel32.DeleteFileW(str(path) + ":Zone.Identifier")
    except Exception:
        pass


def unblock_engine_binaries(engine_exe: Path) -> None:
    """Unblock Engine.exe and sibling binaries Windows may refuse to execute."""
    windows_unblock(engine_exe)
    root = engine_exe.parent
    for pattern in ("*.exe", "*.dll", "*.pyd"):
        for path in root.rglob(pattern):
            windows_unblock(path)


def runtime_work_dir() -> Path:
    """Writable folder for copied inputs + results (avoids OneDrive / Downloads locks)."""
    import os
    import tempfile

    candidates: list[Path] = []
    if getattr(sys, "frozen", False):
        local = Path(os.environ.get("LOCALAPPDATA", str(Path.home() / "AppData" / "Local")))
        candidates.append(local / "SimilarityParser")
    candidates.append(APP_DIR)
    candidates.append(Path(tempfile.gettempdir()) / "SimilarityParser")

    for folder in candidates:
        try:
            folder.mkdir(parents=True, exist_ok=True)
            probe = folder / ".write_test"
            probe.write_text("ok", encoding="utf-8")
            probe.unlink(missing_ok=True)
            return folder
        except OSError:
            continue
    return APP_DIR


def packaged_engine_exe() -> Path:
    return APP_DIR / "Engine" / "Similarity Engine.exe"


def ensure_local_engine(progress_cb=None) -> Path:
    """
    Run Engine from %LOCALAPPDATA%, not from Downloads.

    Windows (and many corporate policies) block launching .exe files that still
    live under Downloads — that shows up as WinError 5 Access is denied.
    """
    import os
    import shutil

    packaged = packaged_engine_exe()
    if not packaged.exists():
        raise FileNotFoundError(f"Engine not found at:\n{packaged}")

    # Dev / non-frozen: no copy needed.
    if not getattr(sys, "frozen", False):
        return packaged

    local_root = Path(os.environ.get("LOCALAPPDATA", str(Path.home() / "AppData" / "Local")))
    dest_dir = local_root / "SimilarityParser" / "Engine"
    dest_exe = dest_dir / "Similarity Engine.exe"
    src_dir = packaged.parent

    def _stamp(folder: Path) -> tuple[int, int]:
        exe = folder / "Similarity Engine.exe"
        if not exe.exists():
            return (0, 0)
        st = exe.stat()
        return (int(st.st_mtime), int(st.st_size))

    needs_copy = _stamp(dest_dir) != _stamp(src_dir)
    if needs_copy:
        if progress_cb:
            progress_cb("Copying engine out of Downloads (one-time)…")
        if dest_dir.exists():
            shutil.rmtree(dest_dir, ignore_errors=True)
        dest_dir.parent.mkdir(parents=True, exist_ok=True)
        shutil.copytree(src_dir, dest_dir)

    if progress_cb:
        progress_cb("Unlocking local engine…")
    unblock_engine_binaries(dest_exe)
    if not dest_exe.exists():
        raise FileNotFoundError(f"Local engine missing after copy:\n{dest_exe}")
    return dest_exe


def similarity_command(engine_exe: Path | None = None) -> list[str]:
    """Command prefix to run the similarity engine (dev script or packaged exe)."""
    if getattr(sys, "frozen", False):
        engine = engine_exe or packaged_engine_exe()
        return [str(engine)]
    return [sys.executable, str(APP_DIR / "similarity.py")]


def access_denied_help(engine: Path | None = None) -> str:
    bits = [
        "Windows blocked running the analysis engine (Access is denied).",
        "",
        "Most common cause: the app is still inside Downloads.",
        "",
        "Fix:",
        "1. Close the app.",
        "2. Move the whole SimilarityParser folder to Desktop (or C:\\Tools).",
        "3. Right-click Start Review.bat → Properties → Unblock (if shown).",
        "4. Run Start Review.bat again.",
        "",
        "Also close FOExport.xlsx in Excel if it is open.",
    ]
    if engine is not None:
        bits.extend(["", f"Engine path:\n{engine}"])
    return "\n".join(bits)

FONT_TITLE = "Bahnschrift"
FONT_MONO = "Cascadia Mono"

# Max semantic "possible related" items shown per cluster in the reviewer panel.
SEMANTIC_MAX_PER_CLUSTER = 20

STATUS_LABELS = {
    "duplicate": "Duplicate",
    "unique": "Unique",
    "discard": "Discard",
    "skip": "Skip",
    "same": "Duplicate",
    "different": "Unique",
}


class Theme:
    """Active UI colors — mutated by apply_theme_palette()."""

    paper = QColor("#E6EAF0")
    paper_deep = QColor("#D5DBE6")
    ink = QColor("#1B2430")
    ink_muted = QColor("#5A6575")
    accent = QColor("#E6A800")
    coral = QColor("#D64545")
    teal = QColor("#0F766E")
    ochre = QColor("#C27803")
    surface = QColor("#F7F8FA")
    # Token chips — high contrast on purpose
    chip_shared_bg = QColor("#FFFFFF")
    chip_shared_fg = QColor("#1B2430")
    chip_shared_border = QColor("#3A4555")
    chip_added_bg = QColor("#0F766E")
    chip_added_fg = QColor("#FFFFFF")
    chip_removed_bg = QColor("#D64545")
    chip_removed_fg = QColor("#FFFFFF")
    # Drawing number matched with parent (xx-xx-xx) — maximum contrast
    chip_drawing_bg = QColor("#FFD000")
    chip_drawing_fg = QColor("#000000")
    chip_drawing_border = QColor("#000000")
    rail_line = QColor("#C5CDD9")
    hover_tint = "#fff8e6"
    dark = False


DRAWING_NUMBER_RE = re.compile(r"^[^\s-]+-[^\s-]+-[^\s-]+$")


def is_drawing_number(token: str) -> bool:
    """True for drawing numbers with exactly two dashes, e.g. 12-34-56 or AB-C1-99."""
    return bool(token) and bool(DRAWING_NUMBER_RE.match(token.strip()))


def apply_theme_palette(*, dark: bool) -> None:
    Theme.dark = dark
    if dark:
        Theme.paper = QColor("#12161C")
        Theme.paper_deep = QColor("#1C222C")
        Theme.ink = QColor("#E8ECF2")
        Theme.ink_muted = QColor("#A8B2C0")
        Theme.accent = QColor("#E6A800")
        Theme.coral = QColor("#FF6B6B")
        Theme.teal = QColor("#2DD4BF")
        Theme.ochre = QColor("#F0B429")
        Theme.surface = QColor("#1A1F28")
        Theme.chip_shared_bg = QColor("#2A3340")
        Theme.chip_shared_fg = QColor("#F2F5FA")
        Theme.chip_shared_border = QColor("#C5CDD9")
        Theme.chip_added_bg = QColor("#0D9488")
        Theme.chip_added_fg = QColor("#041411")
        Theme.chip_removed_bg = QColor("#E11D48")
        Theme.chip_removed_fg = QColor("#FFFFFF")
        Theme.chip_drawing_bg = QColor("#FFD000")
        Theme.chip_drawing_fg = QColor("#000000")
        Theme.chip_drawing_border = QColor("#FFFFFF")
        Theme.rail_line = QColor("#2A3340")
        Theme.hover_tint = "#2A2410"
    else:
        Theme.paper = QColor("#E6EAF0")
        Theme.paper_deep = QColor("#D5DBE6")
        Theme.ink = QColor("#1B2430")
        Theme.ink_muted = QColor("#5A6575")
        Theme.accent = QColor("#E6A800")
        Theme.coral = QColor("#D64545")
        Theme.teal = QColor("#0F766E")
        Theme.ochre = QColor("#C27803")
        Theme.surface = QColor("#F7F8FA")
        Theme.chip_shared_bg = QColor("#FFFFFF")
        Theme.chip_shared_fg = QColor("#1B2430")
        Theme.chip_shared_border = QColor("#3A4555")
        Theme.chip_added_bg = QColor("#0F766E")
        Theme.chip_added_fg = QColor("#FFFFFF")
        Theme.chip_removed_bg = QColor("#D64545")
        Theme.chip_removed_fg = QColor("#FFFFFF")
        Theme.chip_drawing_bg = QColor("#FFD000")
        Theme.chip_drawing_fg = QColor("#000000")
        Theme.chip_drawing_border = QColor("#000000")
        Theme.rail_line = QColor("#C5CDD9")
        Theme.hover_tint = "#fff8e6"


def status_color(status: str) -> QColor:
    return {
        "duplicate": Theme.coral,
        "unique": Theme.teal,
        "discard": Theme.ochre,
        "skip": Theme.ink_muted,
        "unreviewed": Theme.ink_muted,
    }.get(status, Theme.ink_muted)


# Back-compat aliases used widely in paint code
def _sync_legacy_color_aliases() -> None:
    global PAPER, PAPER_DEEP, INK, INK_MUTED, ACCENT, CORAL, TEAL, OCHRE, WHITE, STATUS_COLORS
    PAPER = Theme.paper
    PAPER_DEEP = Theme.paper_deep
    INK = Theme.ink
    INK_MUTED = Theme.ink_muted
    ACCENT = Theme.accent
    CORAL = Theme.coral
    TEAL = Theme.teal
    OCHRE = Theme.ochre
    WHITE = Theme.surface
    STATUS_COLORS = {
        "duplicate": Theme.coral,
        "unique": Theme.teal,
        "discard": Theme.ochre,
        "skip": Theme.ink_muted,
        "unreviewed": Theme.ink_muted,
    }


apply_theme_palette(dark=False)
_sync_legacy_color_aliases()


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def format_duration(seconds: float | int | None) -> str:
    """Human-readable duration for timers and reports."""
    if seconds is None:
        return "—"
    total = max(0, int(round(float(seconds))))
    hours, rem = divmod(total, 3600)
    minutes, secs = divmod(rem, 60)
    if hours:
        return f"{hours}h {minutes}m {secs}s"
    if minutes:
        return f"{minutes}m {secs}s"
    return f"{secs}s"


def parent_times_dict(progress: dict) -> dict[str, float]:
    """Normalized cluster_id → cumulative seconds spent reviewing that parent."""
    raw = progress.get("parent_times") or {}
    out: dict[str, float] = {}
    for key, value in raw.items():
        try:
            secs = float(value or 0)
        except (TypeError, ValueError):
            continue
        if secs <= 0:
            continue
        out[str(key)] = secs
    return out


def clean_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value != value:  # NaN
        return ""
    return str(value).strip()


def clean_description(value) -> str:
    return clean_text(value).lstrip()


def normalize_status(status: str) -> str:
    if status in ("same", "duplicate"):
        return "duplicate"
    if status in ("different", "unique"):
        return "unique"
    return status


def tokenize(text: str) -> frozenset[str]:
    """Same bag-of-words normalization as similarity.py / text_normalize."""
    return frozenset(normalize_tokens(text))


def token_diff(ref_text: str, cand_text: str) -> tuple[set[str], set[str], set[str]]:
    """Return (shared, only_in_ref, only_in_cand) token sets."""
    ref = tokenize(ref_text)
    cand = tokenize(cand_text)
    return ref & cand, ref - cand, cand - ref


def format_score(score) -> str:
    if isinstance(score, (int, float)) and score == score:
        return f"{score:.2f}"
    return "—"


def score_value(score) -> float:
    if isinstance(score, (int, float)) and score == score:
        return float(score)
    return 0.0


def pick_font(preferred: str, fallback: str, point_size: int, weight: int = QFont.Weight.Normal) -> QFont:
    font = QFont(preferred, point_size, weight)
    if font.exactMatch() or QFont(preferred).family() == preferred:
        return font
    return QFont(fallback, point_size, weight)


def _safe_int(value, default: int = 0) -> int:
    try:
        if value is None or (isinstance(value, float) and value != value):
            return default
        return int(value)
    except (TypeError, ValueError):
        return default


def workbook_has_sheet(path: Path, sheet_name: str) -> bool:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        return sheet_name in wb.sheetnames
    finally:
        wb.close()


def load_grouped_review(path: Path) -> tuple[list[int], dict[int, list[dict]], dict[str, dict]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if "Grouped Review" not in wb.sheetnames:
            raise ValueError("Workbook has no 'Grouped Review' sheet.")
        ws = wb["Grouped Review"]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return [], {}, {}
        headers = [clean_text(h) for h in header_row]
        col = {name: i for i, name in enumerate(headers)}
        required = (
            "cluster_id", "cluster_size", "position_in_cluster", "depth",
            "product_number", "description", "linked_to_product",
            "score_to_parent", "n_similar_in_cluster", "exact_dup_group",
        )
        missing = [name for name in required if name not in col]
        if missing:
            raise ValueError(f"Grouped Review missing columns: {missing}")

        cluster_order: list[int] = []
        clusters: dict[int, list[dict]] = {}
        by_product: dict[str, dict] = {}

        for row in rows_iter:
            if row is None or all(v is None or v == "" for v in row):
                continue

            def cell(name: str):
                return row[col[name]] if col[name] < len(row) else None

            item = {
                "cluster_id": _safe_int(cell("cluster_id")),
                "cluster_size": _safe_int(cell("cluster_size")),
                "position_in_cluster": _safe_int(cell("position_in_cluster")),
                "depth": _safe_int(cell("depth")),
                "product_number": clean_text(cell("product_number")),
                "description": clean_description(cell("description")),
                "linked_to_product": clean_text(cell("linked_to_product")),
                "score_to_parent": cell("score_to_parent"),
                "n_similar_in_cluster": _safe_int(cell("n_similar_in_cluster")),
                "exact_dup_group": cell("exact_dup_group"),
            }
            if not item["product_number"]:
                continue
            cid = item["cluster_id"]
            if cid not in clusters:
                cluster_order.append(cid)
                clusters[cid] = []
            clusters[cid].append(item)
            by_product[item["product_number"]] = item
    finally:
        wb.close()

    return cluster_order, clusters, by_product


def _refresh_cluster_sizes(clusters: dict[int, list[dict]]) -> None:
    for members in clusters.values():
        size = len(members)
        for it in members:
            it["cluster_size"] = size
            it["n_similar_in_cluster"] = max(size - 1, 0)


def apply_cluster_moves(
    cluster_order: list[int],
    clusters: dict[int, list[dict]],
    by_product: dict[str, dict],
    moves: dict,
) -> list[int]:
    """Re-apply reviewer pulls from Related into clusters (sidecar only)."""
    if not moves:
        return cluster_order

    for pn, move in moves.items():
        item = by_product.get(pn)
        if item is None:
            continue
        try:
            new_cid = int(move.get("cluster_id"))
        except (TypeError, ValueError):
            continue
        old_cid = item.get("cluster_id")
        if old_cid == new_cid:
            # Still refresh link fields if present
            linked = move.get("linked_to_product")
            if linked:
                item["linked_to_product"] = linked
            continue

        if old_cid in clusters:
            clusters[old_cid] = [
                i for i in clusters[old_cid] if i["product_number"] != pn
            ]
            if not clusters[old_cid]:
                del clusters[old_cid]

        item["cluster_id"] = new_cid
        linked = move.get("linked_to_product") or ""
        if linked:
            item["linked_to_product"] = linked
        if item.get("depth", 0) == 0:
            item["depth"] = 1
        score = move.get("semantic_score")
        if score is not None and score != "":
            try:
                item["score_to_parent"] = float(score)
            except (TypeError, ValueError):
                pass

        clusters.setdefault(new_cid, [])
        if not any(i["product_number"] == pn for i in clusters[new_cid]):
            clusters[new_cid].append(item)
        if new_cid not in cluster_order:
            cluster_order.append(new_cid)

    # Drop empty clusters from navigation order
    cluster_order = [cid for cid in cluster_order if clusters.get(cid)]
    _refresh_cluster_sizes(clusters)
    return cluster_order


def load_semantic_suggestions(path: Path) -> dict[str, list[dict]]:
    """Read the optional 'Semantic Suggestions' sheet.

    Returns product_number -> list of suggestion dicts (sorted by score desc).
    Missing sheet or any read error yields an empty mapping so older workbooks
    still open cleanly.
    """
    from openpyxl import load_workbook

    out: dict[str, list[dict]] = {}
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return out
    try:
        if "Semantic Suggestions" not in wb.sheetnames:
            return out
        ws = wb["Semantic Suggestions"]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return out
        col = {clean_text(h): i for i, h in enumerate(header_row)}
        required = ("product_number", "suggested_product")
        if any(name not in col for name in required):
            return out

        def cell(row, name):
            idx = col.get(name)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        for row in rows_iter:
            if row is None or all(v is None or v == "" for v in row):
                continue
            pn = clean_text(cell(row, "product_number"))
            suggested = clean_text(cell(row, "suggested_product"))
            if not pn or not suggested:
                continue
            out.setdefault(pn, []).append({
                "suggested_product": suggested,
                "suggested_cluster_id": _safe_int(cell(row, "suggested_cluster_id")),
                "suggested_description": clean_description(cell(row, "suggested_description")),
                "semantic_score": cell(row, "semantic_score"),
            })
    finally:
        wb.close()

    for pn in out:
        out[pn].sort(key=lambda s: score_value(s["semantic_score"]), reverse=True)
    return out


def progress_path_for(results_path: Path) -> Path:
    """One progress file per results workbook so different inputs stay separate."""
    safe = results_path.stem.replace(" ", "_")
    return results_path.parent / f"review_progress__{safe}.json"


def decisions_path_for(results_path: Path) -> Path:
    """Human-readable decisions Excel — separate from the original results workbook."""
    safe = results_path.stem.replace(" ", "_")
    return results_path.parent / f"review_decisions__{safe}.xlsx"


def load_progress(results_path: Path | None = None) -> dict:
    path = progress_path_for(results_path) if results_path else PROGRESS_FILE
    if not path.exists() and results_path is not None and PROGRESS_FILE.exists():
        # Migrate legacy single progress file when it matches this source
        with PROGRESS_FILE.open(encoding="utf-8") as f:
            legacy = json.load(f)
        if legacy.get("source_file") in ("", results_path.name):
            path = PROGRESS_FILE
    if not path.exists():
        return {
            "source_file": results_path.name if results_path else "",
            "decisions": {},
            "clusters_completed": [],
            "parent_times": {},
            "cluster_moves": {},
        }
    with path.open(encoding="utf-8") as f:
        data = json.load(f)
    data.setdefault("decisions", {})
    data.setdefault("clusters_completed", [])
    data.setdefault("parent_times", {})
    data.setdefault("cluster_moves", {})
    if results_path is not None:
        data["source_file"] = results_path.name
    return data


def save_progress(progress: dict, results_path: Path | None = None) -> None:
    """Write JSON progress only. Never touches the original results workbook."""
    progress["updated_at"] = utc_now()
    if results_path is not None:
        path = progress_path_for(results_path)
        progress["source_file"] = results_path.name
    else:
        path = PROGRESS_FILE
    with path.open("w", encoding="utf-8") as f:
        json.dump(progress, f, indent=2)


def save_decisions_workbook(
    results_path: Path,
    by_product: dict[str, dict],
    progress: dict,
) -> Path:
    """Write a companion Excel of decisions; leaves the original results file alone."""
    from openpyxl import Workbook

    out = decisions_path_for(results_path)
    decisions = progress.get("decisions", {})
    rows: list[list] = []
    for pn, dec in decisions.items():
        item = by_product.get(pn, {})
        status = normalize_status(dec.get("status", ""))
        note = dec.get("note", "")
        linked = item.get("linked_to_product", "")
        rows.append([
            pn,
            STATUS_LABELS.get(status, status),
            dec.get("cluster_id", item.get("cluster_id", "")),
            item.get("description", ""),
            linked,
            item.get("score_to_parent", ""),
            note,
            dec.get("updated_at", ""),
        ])
    rows.sort(key=lambda r: (str(r[1]), str(r[2]), str(r[0])))

    wb = Workbook()
    ws = wb.active
    ws.title = "Decisions"
    ws.append([
        "Product", "Decision", "Cluster", "Description",
        "Linked to", "Score to parent", "Note", "Updated (UTC)",
    ])
    for row in rows:
        ws.append(row)

    moves = progress.get("cluster_moves", {}) or {}
    if moves:
        ms = wb.create_sheet("Cluster Moves")
        ms.append([
            "Product", "To cluster", "From cluster", "Linked to", "Updated (UTC)",
        ])
        for pn, move in sorted(moves.items()):
            ms.append([
                pn,
                move.get("cluster_id", ""),
                move.get("from_cluster_id", ""),
                move.get("linked_to_product", ""),
                move.get("updated_at", ""),
            ])

    meta = wb.create_sheet("About", 0)
    meta.append(["Field", "Value"])
    meta.append(["Source results workbook", results_path.name])
    meta.append(["Source path", str(results_path)])
    meta.append(["Decisions file", out.name])
    meta.append(["Updated (UTC)", progress.get("updated_at", utc_now())])
    meta.append(["Note", "Original results workbook is never modified by the reviewer."])
    meta.append(["Decision count", len(rows)])
    meta.append(["Cluster moves (Related pulls)", len(moves)])

    wb.save(out)
    return out


def compute_review_stats(
    by_product: dict[str, dict],
    cluster_order: list[int],
    progress: dict,
    *,
    current_cluster_id: int | None = None,
    current_cluster_index: int = 0,
) -> dict:
    decisions = progress.get("decisions", {})
    completed = set(progress.get("clusters_completed", []))

    status_counts: Counter[str] = Counter()
    for pn, dec in decisions.items():
        if pn not in by_product:
            continue
        status_counts[normalize_status(dec["status"])] += 1

    reviewed = sum(status_counts.values())
    total = len(by_product)
    remaining = max(total - reviewed, 0)
    n_clusters = len(cluster_order)
    n_clusters_done = sum(1 for cid in cluster_order if cid in completed)

    dup = status_counts["duplicate"]
    uniq = status_counts["unique"]
    discard = status_counts["discard"]
    skip = status_counts.get("skip", 0)

    def pct(part: int, whole: int) -> float:
        return (100.0 * part / whole) if whole else 0.0

    current_cluster_reviewed = 0
    current_cluster_size = 0
    if current_cluster_id is not None and current_cluster_id in cluster_order:
        current_cluster_size = sum(
            1 for item in by_product.values()
            if item["cluster_id"] == current_cluster_id
        )
        current_cluster_reviewed = sum(
            1 for pn in decisions
            if pn in by_product and by_product[pn]["cluster_id"] == current_cluster_id
        )

    times = parent_times_dict(progress)
    timed_parents = len(times)
    total_parent_seconds = sum(times.values())
    avg_seconds_per_parent = (
        total_parent_seconds / timed_parents if timed_parents else 0.0
    )

    return {
        "total": total,
        "reviewed": reviewed,
        "remaining": remaining,
        "reviewed_pct": pct(reviewed, total),
        "n_clusters": n_clusters,
        "n_clusters_done": n_clusters_done,
        "clusters_done_pct": pct(n_clusters_done, n_clusters),
        "duplicate": dup,
        "unique": uniq,
        "discard": discard,
        "skip": skip,
        "duplicate_pct_of_reviewed": pct(dup, reviewed),
        "unique_pct_of_reviewed": pct(uniq, reviewed),
        "discard_pct_of_reviewed": pct(discard, reviewed),
        "skip_pct_of_reviewed": pct(skip, reviewed),
        "duplicate_pct_of_total": pct(dup, total),
        "current_cluster_id": current_cluster_id,
        "current_cluster_index": current_cluster_index,
        "current_cluster_reviewed": current_cluster_reviewed,
        "current_cluster_size": current_cluster_size,
        "timed_parents": timed_parents,
        "total_parent_seconds": total_parent_seconds,
        "avg_seconds_per_parent": avg_seconds_per_parent,
        "source_file": progress.get("source_file", ""),
        "updated_at": progress.get("updated_at", ""),
    }


def format_review_stats(stats: dict) -> str:
    lines = [
        "REVIEW STATISTICS",
        "=" * 40,
        "",
        "Overall progress",
        f"  Reviewed     {stats['reviewed']:,} / {stats['total']:,}  ({stats['reviewed_pct']:.1f}%)",
        f"  Remaining    {stats['remaining']:,}",
        "",
        "Clusters",
        f"  Completed    {stats['n_clusters_done']:,} / {stats['n_clusters']:,}  ({stats['clusters_done_pct']:.1f}%)",
    ]
    if stats["current_cluster_id"] is not None:
        lines.extend([
            f"  Current      cluster {stats['current_cluster_index'] + 1} "
            f"(id {stats['current_cluster_id']}) — "
            f"{stats['current_cluster_reviewed']:,} / {stats['current_cluster_size']:,} marked",
        ])
    lines.extend([
        "",
        "Time per parent",
        f"  Parents timed {stats['timed_parents']:,}",
        f"  Total time    {format_duration(stats['total_parent_seconds'])}",
        f"  Average       {format_duration(stats['avg_seconds_per_parent'])} per parent",
    ])
    lines.extend([
        "",
        "Decisions (reviewed items only)",
        f"  Duplicate    {stats['duplicate']:,}  ({stats['duplicate_pct_of_reviewed']:.1f}%)",
        f"  Unique       {stats['unique']:,}  ({stats['unique_pct_of_reviewed']:.1f}%)",
        f"  Discard      {stats['discard']:,}  ({stats['discard_pct_of_reviewed']:.1f}%)",
    ])
    if stats["skip"]:
        lines.append(
            f"  Skip         {stats['skip']:,}  ({stats['skip_pct_of_reviewed']:.1f}%)",
        )
    lines.extend([
        "",
        "Key rates",
        f"  Duplicate rate   {stats['duplicate_pct_of_reviewed']:.1f}% of reviewed",
        f"  Unique rate      {stats['unique_pct_of_reviewed']:.1f}% of reviewed",
        f"  Duplicates found {stats['duplicate']:,} ({stats['duplicate_pct_of_total']:.1f}% of all products)",
        "",
        "Session",
        f"  Source file  {stats['source_file'] or '(none)'}",
        f"  Last saved   {stats['updated_at'] or '(never)'}",
    ])
    return "\n".join(lines)


def build_decision_rows(
    by_product: dict[str, dict],
    progress: dict,
) -> list[dict]:
    """Flat decision rows for reporting (product, status, cluster, linked parent)."""
    decisions = progress.get("decisions", {})
    rows: list[dict] = []
    for pn, dec in decisions.items():
        item = by_product.get(pn)
        if item is None:
            continue
        status = normalize_status(dec["status"])
        rows.append({
            "product_number": pn,
            "status": status,
            "status_label": STATUS_LABELS.get(status, status),
            "cluster_id": item["cluster_id"],
            "description": item["description"],
            "linked_to_product": item["linked_to_product"],
            "score_to_parent": item["score_to_parent"],
            "updated_at": dec.get("updated_at", ""),
        })
    rows.sort(key=lambda r: (r["status"], r["cluster_id"], r["product_number"]))
    return rows


def build_cluster_report_rows(
    cluster_order: list[int],
    clusters: dict[int, list[dict]],
    progress: dict,
) -> list[dict]:
    decisions = progress.get("decisions", {})
    completed = set(progress.get("clusters_completed", []))
    times = parent_times_dict(progress)
    rows: list[dict] = []
    for cid in cluster_order:
        items = clusters.get(cid, [])
        counts: Counter[str] = Counter()
        for item in items:
            pn = item["product_number"]
            if pn in decisions:
                counts[normalize_status(decisions[pn]["status"])] += 1
        marked = sum(counts.values())
        size = len(items)
        root = cluster_root(items)
        seconds = times.get(str(cid), 0.0)
        rows.append({
            "cluster_id": cid,
            "size": size,
            "marked": marked,
            "remaining": max(size - marked, 0),
            "duplicate": counts["duplicate"],
            "unique": counts["unique"],
            "discard": counts["discard"],
            "completed": cid in completed,
            "reference": root["product_number"] if root else "",
            "time_seconds": seconds,
            "time_label": format_duration(seconds) if seconds else "—",
        })
    return rows


def export_management_report(
    path: Path,
    *,
    stats: dict,
    decision_rows: list[dict],
    cluster_rows: list[dict],
) -> None:
    """Write a superior-facing Excel workbook."""
    from openpyxl import Workbook

    wb = Workbook()

    def write_sheet(title: str, headers: list[str], rows: list[list]) -> None:
        if title == "Executive Summary" and wb.active and wb.active.title == "Sheet":
            ws = wb.active
            ws.title = title
        else:
            ws = wb.create_sheet(title)
        ws.append(headers)
        for row in rows:
            ws.append(row)

    write_sheet(
        "Executive Summary",
        ["Metric", "Value"],
        [
            ["Source file", stats.get("source_file") or ""],
            ["Report generated (UTC)", utc_now()],
            ["Last progress save", stats.get("updated_at") or ""],
            ["Products total", stats["total"]],
            ["Reviewed", stats["reviewed"]],
            ["Remaining", stats["remaining"]],
            ["Reviewed %", round(stats["reviewed_pct"], 1)],
            ["Clusters total", stats["n_clusters"]],
            ["Clusters completed", stats["n_clusters_done"]],
            ["Clusters completed %", round(stats["clusters_done_pct"], 1)],
            ["Parents with time recorded", stats["timed_parents"]],
            ["Total time on parents", format_duration(stats["total_parent_seconds"])],
            ["Average time per parent", format_duration(stats["avg_seconds_per_parent"])],
            ["Average time per parent (seconds)", round(stats["avg_seconds_per_parent"], 1)],
            ["Marked duplicate", stats["duplicate"]],
            ["Marked unique", stats["unique"]],
            ["Marked discard", stats["discard"]],
            ["Duplicate % of reviewed", round(stats["duplicate_pct_of_reviewed"], 1)],
            ["Duplicate % of all products", round(stats["duplicate_pct_of_total"], 1)],
        ],
    )

    write_sheet(
        "Duplicates",
        ["Product", "Cluster", "Description", "Linked to", "Score to parent", "Updated (UTC)"],
        [
            [
                r["product_number"], r["cluster_id"], r["description"],
                r["linked_to_product"], r["score_to_parent"], r["updated_at"],
            ]
            for r in decision_rows if r["status"] == "duplicate"
        ],
    )

    write_sheet(
        "All Decisions",
        ["Product", "Decision", "Cluster", "Description", "Linked to", "Score to parent", "Updated (UTC)"],
        [
            [
                r["product_number"], r["status_label"], r["cluster_id"], r["description"],
                r["linked_to_product"], r["score_to_parent"], r["updated_at"],
            ]
            for r in decision_rows
        ],
    )

    write_sheet(
        "Cluster Progress",
        [
            "Cluster", "Reference", "Size", "Marked", "Remaining",
            "Duplicate", "Unique", "Discard", "Completed", "Time", "Time (seconds)",
        ],
        [
            [
                r["cluster_id"], r["reference"], r["size"], r["marked"], r["remaining"],
                r["duplicate"], r["unique"], r["discard"], r["completed"],
                r["time_label"], round(r["time_seconds"], 1),
            ]
            for r in cluster_rows
        ],
    )

    wb.save(path)


def _score_sort_key(item: dict) -> tuple:
    return (-score_value(item["score_to_parent"]), item["product_number"])


def cluster_root(items: list[dict]) -> dict | None:
    if not items:
        return None
    roots = [item for item in items if item["depth"] == 0]
    if not roots:
        return min(items, key=lambda i: i["position_in_cluster"])
    return min(roots, key=lambda i: i["position_in_cluster"])


class ReportsScreen(QWidget):
    """Management-facing report view with Excel export."""

    def __init__(self, window: "ReviewWindow") -> None:
        super().__init__()
        self._window = window
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 16, 20, 16)
        layout.setSpacing(12)

        header = QHBoxLayout()
        title = QLabel("REPORTS")
        title.setFont(pick_font(FONT_TITLE, "Segoe UI", 18, QFont.Weight.Bold))
        self.title = title
        header.addWidget(title)
        header.addStretch(1)

        self.refresh_btn = QPushButton("Refresh")
        self.refresh_btn.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.refresh_btn.clicked.connect(self.refresh)
        header.addWidget(self.refresh_btn)

        self.export_btn = QPushButton("Export Excel…")
        self.export_btn.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.export_btn.clicked.connect(self._export_excel)
        header.addWidget(self.export_btn)

        self.back_btn = QPushButton("← Back to review")
        self.back_btn.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.back_btn.clicked.connect(window.show_review_screen)
        header.addWidget(self.back_btn)
        layout.addLayout(header)

        self.kpi_row = QHBoxLayout()
        self.kpi_labels: list[QLabel] = []
        for _ in range(4):
            lab = QLabel("")
            lab.setWordWrap(True)
            lab.setMinimumHeight(72)
            self.kpi_labels.append(lab)
            self.kpi_row.addWidget(lab, 1)
        layout.addLayout(self.kpi_row)

        self.summary = QTextEdit()
        self.summary.setReadOnly(True)
        self.summary.setMaximumHeight(160)
        layout.addWidget(self.summary)

        split = QHBoxLayout()
        left = QVBoxLayout()
        self.dup_caption = QLabel("Duplicates marked (for superiors)")
        self.dup_caption.setFont(pick_font(FONT_TITLE, "Segoe UI", 11, QFont.Weight.Bold))
        left.addWidget(self.dup_caption)
        self.dup_table = QTableWidget(0, 5)
        self.dup_table.setHorizontalHeaderLabels(
            ["Product", "Cluster", "Linked to", "Score", "Updated"],
        )
        self.dup_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.dup_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.dup_table.horizontalHeader().setStretchLastSection(True)
        left.addWidget(self.dup_table, 1)

        right = QVBoxLayout()
        self.cluster_caption = QLabel("Cluster progress")
        self.cluster_caption.setFont(pick_font(FONT_TITLE, "Segoe UI", 11, QFont.Weight.Bold))
        right.addWidget(self.cluster_caption)
        self.cluster_table = QTableWidget(0, 8)
        self.cluster_table.setHorizontalHeaderLabels(
            ["Cluster", "Reference", "Size", "Marked", "Dup", "Unique", "Done", "Time"],
        )
        self.cluster_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.cluster_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.cluster_table.horizontalHeader().setStretchLastSection(True)
        right.addWidget(self.cluster_table, 1)

        split.addLayout(left, 3)
        split.addLayout(right, 2)
        layout.addLayout(split, 1)

        self._decision_rows: list[dict] = []
        self._cluster_rows: list[dict] = []
        self._stats: dict = {}
        self.apply_theme()

    def apply_theme(self) -> None:
        self.title.setStyleSheet(f"color: {Theme.ink.name()}; letter-spacing: 2px;")
        self.dup_caption.setStyleSheet(f"color: {Theme.ink.name()};")
        self.cluster_caption.setStyleSheet(f"color: {Theme.ink.name()};")
        btn_css = f"""
            QPushButton {{
                background: {Theme.surface.name()};
                color: {Theme.ink.name()};
                border: 1px solid {Theme.paper_deep.name()};
                border-radius: 2px;
                padding: 8px 14px;
                font-family: {FONT_TITLE};
                font-size: 10pt;
            }}
            QPushButton:hover {{
                border-color: {Theme.accent.name()};
                background: {Theme.hover_tint};
            }}
        """
        for btn in (self.refresh_btn, self.export_btn, self.back_btn):
            btn.setStyleSheet(btn_css)
        self.summary.setStyleSheet(
            f"background: {Theme.surface.name()}; color: {Theme.ink.name()}; "
            f"border: 1px solid {Theme.paper_deep.name()}; "
            f"font-family: Consolas, 'Cascadia Mono', monospace; font-size: 10pt;",
        )
        table_css = f"""
            QTableWidget {{
                background: {Theme.surface.name()};
                color: {Theme.ink.name()};
                gridline-color: {Theme.paper_deep.name()};
                border: 1px solid {Theme.paper_deep.name()};
            }}
            QHeaderView::section {{
                background: {Theme.paper_deep.name()};
                color: {Theme.ink.name()};
                padding: 6px;
                border: none;
                font-weight: 700;
            }}
        """
        self.dup_table.setStyleSheet(table_css)
        self.cluster_table.setStyleSheet(table_css)
        for lab in self.kpi_labels:
            lab.setStyleSheet(
                f"background: {Theme.surface.name()}; color: {Theme.ink.name()}; "
                f"border: 1px solid {Theme.paper_deep.name()}; padding: 10px; "
                f"font-family: {FONT_TITLE}; font-size: 11pt;",
            )

    def refresh(self) -> None:
        w = self._window
        self._stats = compute_review_stats(
            w.by_product,
            w.cluster_order,
            w.progress,
            current_cluster_id=w._current_cluster_id(),
            current_cluster_index=w.cluster_index,
        )
        self._decision_rows = build_decision_rows(w.by_product, w.progress)
        self._cluster_rows = build_cluster_report_rows(
            w.cluster_order, w.clusters, w.progress,
        )
        s = self._stats
        kpi_texts = [
            f"Reviewed\n{s['reviewed']:,} / {s['total']:,}\n{s['reviewed_pct']:.1f}%",
            f"Duplicates\n{s['duplicate']:,}\n{s['duplicate_pct_of_total']:.1f}% of all",
            f"Unique\n{s['unique']:,}\n{s['unique_pct_of_reviewed']:.1f}% of reviewed",
            f"Avg / parent\n{format_duration(s['avg_seconds_per_parent'])}\n"
            f"{s['timed_parents']:,} timed · {format_duration(s['total_parent_seconds'])} total",
        ]
        for lab, text in zip(self.kpi_labels, kpi_texts):
            lab.setText(text)
        self.summary.setPlainText(format_review_stats(s))

        dups = [r for r in self._decision_rows if r["status"] == "duplicate"]
        self.dup_table.setRowCount(len(dups))
        for i, row in enumerate(dups):
            values = [
                row["product_number"],
                str(row["cluster_id"]),
                row["linked_to_product"],
                format_score(row["score_to_parent"]),
                row["updated_at"],
            ]
            for j, val in enumerate(values):
                self.dup_table.setItem(i, j, QTableWidgetItem(val))

        self.cluster_table.setRowCount(len(self._cluster_rows))
        for i, row in enumerate(self._cluster_rows):
            values = [
                str(row["cluster_id"]),
                row["reference"],
                str(row["size"]),
                str(row["marked"]),
                str(row["duplicate"]),
                str(row["unique"]),
                "Yes" if row["completed"] else "No",
                row["time_label"],
            ]
            for j, val in enumerate(values):
                self.cluster_table.setItem(i, j, QTableWidgetItem(val))

    def _export_excel(self) -> None:
        from PySide6.QtWidgets import QFileDialog

        self.refresh()
        default = APP_DIR / f"similarity_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Export management report",
            str(default),
            "Excel files (*.xlsx)",
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        try:
            export_management_report(
                Path(path),
                stats=self._stats,
                decision_rows=self._decision_rows,
                cluster_rows=self._cluster_rows,
            )
        except Exception as exc:
            QMessageBox.critical(self, "Export failed", str(exc))
            return
        QMessageBox.information(self, "Export complete", f"Saved report to:\n{path}")


class HatchPaper(QWidget):
    """Cool paper background with a subtle diagonal hatch."""

    def paintEvent(self, _event) -> None:
        painter = QPainter(self)
        painter.fillRect(self.rect(), Theme.paper)
        painter.setPen(QPen(Theme.paper_deep, 1))
        step = 14
        w, h = self.width(), self.height()
        for i in range(-h, w + h, step):
            painter.drawLine(i, 0, i + h, h)


class StatsDialog(QDialog):
    def __init__(self, window: "ReviewWindow") -> None:
        super().__init__(window)
        self.setWindowTitle("Review statistics")
        self.setMinimumSize(480, 420)
        self.setStyleSheet(f"""
            QDialog {{
                background-color: {Theme.paper.name()};
                color: {Theme.ink.name()};
            }}
            QTextEdit {{
                background-color: {Theme.surface.name()};
                color: {Theme.ink.name()};
                border: 1px solid {Theme.paper_deep.name()};
                font-family: Consolas, "Cascadia Mono", monospace;
                font-size: 11pt;
            }}
            QPushButton {{
                background-color: {Theme.surface.name()};
                color: {Theme.ink.name()};
                border: 1px solid {Theme.ink_muted.name()};
                border-radius: 2px;
                padding: 8px 16px;
                font-family: {FONT_TITLE};
            }}
            QPushButton:hover {{
                border-color: {Theme.accent.name()};
            }}
        """)
        layout = QVBoxLayout(self)
        self.text = QTextEdit()
        self.text.setReadOnly(True)
        layout.addWidget(self.text)
        refresh_btn = QPushButton("Refresh")
        refresh_btn.clicked.connect(lambda: self._refresh(window))
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(self.accept)
        btn_row = QHBoxLayout()
        btn_row.addStretch(1)
        btn_row.addWidget(refresh_btn)
        btn_row.addWidget(close_btn)
        layout.addLayout(btn_row)
        self._refresh(window)

    def _refresh(self, window: "ReviewWindow") -> None:
        stats = compute_review_stats(
            window.by_product,
            window.cluster_order,
            window.progress,
            current_cluster_id=window._current_cluster_id(),
            current_cluster_index=window.cluster_index,
        )
        self.text.setPlainText(format_review_stats(stats))


class TopStrip(QWidget):
    """Status header only — brand + cluster/progress info."""

    def __init__(self, window: "ReviewWindow") -> None:
        super().__init__()
        self._window = window
        self.setFixedHeight(52)
        layout = QHBoxLayout(self)
        layout.setContentsMargins(16, 8, 16, 8)
        layout.setSpacing(12)

        self.brand = QLabel("SIMILARITY")
        self.brand.setFont(pick_font(FONT_TITLE, "Segoe UI", 16, QFont.Weight.Bold))
        layout.addWidget(self.brand)

        self.cluster_label = QLabel("")
        self.cluster_label.setFont(pick_font(FONT_TITLE, "Segoe UI", 11))
        layout.addWidget(self.cluster_label, 1)

        self.progress_label = QLabel("")
        self.progress_label.setFont(pick_font(FONT_MONO, "Consolas", 10))
        layout.addWidget(self.progress_label)
        self.apply_theme()

    def paintEvent(self, event) -> None:
        painter = QPainter(self)
        painter.fillRect(self.rect(), Theme.surface)
        painter.setPen(QPen(Theme.paper_deep, 1))
        painter.drawLine(0, self.height() - 1, self.width(), self.height() - 1)
        super().paintEvent(event)

    def set_info(self, cluster_text: str, progress_text: str) -> None:
        self.cluster_label.setText(cluster_text)
        self.progress_label.setText(progress_text)

    def apply_theme(self) -> None:
        self.brand.setStyleSheet(f"color: {Theme.ink.name()}; letter-spacing: 2px;")
        self.cluster_label.setStyleSheet(f"color: {Theme.ink_muted.name()};")
        self.progress_label.setStyleSheet(f"color: {Theme.ink.name()};")
        self.update()


class TabStrip(QWidget):
    """Tab / action row directly under the top bar."""

    def __init__(self, window: "ReviewWindow") -> None:
        super().__init__()
        self._window = window
        self.setFixedHeight(48)
        layout = QHBoxLayout(self)
        layout.setContentsMargins(12, 6, 12, 6)
        layout.setSpacing(8)

        self._tab_buttons: list[QPushButton] = []
        self._action_buttons: list[QPushButton] = []

        for text, slot in (
            ("Review", window.show_review_screen),
            ("Reports", window.show_reports_screen),
        ):
            btn = QPushButton(text)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setFocusPolicy(Qt.FocusPolicy.NoFocus)
            btn.clicked.connect(slot)
            layout.addWidget(btn)
            self._tab_buttons.append(btn)

        layout.addSpacing(16)

        for text, slot in (
            ("Dark", window._toggle_dark_mode),
            ("Related", window._toggle_related_panel),
            ("Results…", window._pick_results),
            ("Input…", window._pick_and_run_input),
            ("◀ Prev parent", window._prev_cluster),
            ("Next parent →", window._finish_cluster_and_next),
        ):
            btn = QPushButton(text)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setFocusPolicy(Qt.FocusPolicy.NoFocus)
            btn.clicked.connect(slot)
            layout.addWidget(btn)
            self._action_buttons.append(btn)

        layout.addStretch(1)

        self.search_field = QLineEdit()
        self.search_field.setPlaceholderText("Go to parent code…")
        self.search_field.setClearButtonEnabled(True)
        self.search_field.setFixedWidth(190)
        self.search_field.returnPressed.connect(self._on_search)
        layout.addWidget(self.search_field)

        self.search_btn = QPushButton("Go")
        self.search_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.search_btn.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.search_btn.clicked.connect(self._on_search)
        layout.addWidget(self.search_btn)

        self.review_btn = self._tab_buttons[0]
        self.reports_btn = self._tab_buttons[1]
        self.dark_btn = self._action_buttons[0]
        self.apply_theme()

    def _on_search(self) -> None:
        self._window._search_parent(self.search_field.text())

    def focus_search(self) -> None:
        self.search_field.setFocus(Qt.FocusReason.ShortcutFocusReason)
        self.search_field.selectAll()

    def paintEvent(self, event) -> None:
        painter = QPainter(self)
        painter.fillRect(self.rect(), Theme.paper_deep)
        painter.setPen(QPen(Theme.rail_line, 1))
        painter.drawLine(0, self.height() - 1, self.width(), self.height() - 1)
        super().paintEvent(event)

    def apply_theme(self) -> None:
        self.dark_btn.setText("Light" if Theme.dark else "Dark")
        btn_css = f"""
            QPushButton {{
                background: {Theme.surface.name()};
                color: {Theme.ink.name()};
                border: 1px solid {Theme.rail_line.name()};
                border-radius: 2px;
                padding: 7px 14px;
                font-family: {FONT_TITLE};
                font-size: 10pt;
            }}
            QPushButton:hover {{
                border-color: {Theme.accent.name()};
                background: {Theme.hover_tint};
            }}
        """
        active_css = f"""
            QPushButton {{
                background: {Theme.accent.name()};
                color: #1B2430;
                border: 1px solid {Theme.accent.name()};
                border-radius: 2px;
                padding: 7px 18px;
                font-family: {FONT_TITLE};
                font-size: 10pt;
                font-weight: 700;
            }}
        """
        for btn in self._tab_buttons + self._action_buttons:
            btn.setStyleSheet(btn_css)
        self.search_btn.setStyleSheet(btn_css)
        self.search_field.setStyleSheet(f"""
            QLineEdit {{
                background: {Theme.surface.name()};
                color: {Theme.ink.name()};
                border: 1px solid {Theme.rail_line.name()};
                border-radius: 2px;
                padding: 6px 8px;
                font-family: {FONT_MONO};
                font-size: 10pt;
            }}
            QLineEdit:focus {{
                border-color: {Theme.accent.name()};
            }}
        """)
        if getattr(self._window, "_screen", "review") == "reports":
            self.reports_btn.setStyleSheet(active_css)
        else:
            self.review_btn.setStyleSheet(active_css)
        self.update()


class QueueRail(QWidget):
    """Left queue rail: product stubs + status dots."""

    STUB_H = 52
    RAIL_W = 120

    def __init__(self, window: "ReviewWindow") -> None:
        super().__init__()
        self._window = window
        self.setFixedWidth(self.RAIL_W)
        self._order: list[str] = []
        self._statuses: dict[str, str] = {}
        self._selected = ""
        self._scroll = 0
        self.setMouseTracking(True)
        self.setFocusPolicy(Qt.FocusPolicy.NoFocus)

    def set_queue(self, order: list[str], statuses: dict[str, str], selected: str) -> None:
        self._order = list(order)
        self._statuses = dict(statuses)
        self._selected = selected
        self._ensure_selected_visible()
        self.update()

    def set_selected(self, product_number: str) -> None:
        self._selected = product_number
        self._ensure_selected_visible()
        self.update()

    def update_status(self, product_number: str, status: str) -> None:
        self._statuses[product_number] = status
        self.update()

    def _ensure_selected_visible(self) -> None:
        if not self._order or self._selected not in self._order:
            return
        idx = self._order.index(self._selected)
        y = idx * self.STUB_H
        view_h = max(self.height() - 8, self.STUB_H)
        if y < self._scroll:
            self._scroll = y
        elif y + self.STUB_H > self._scroll + view_h:
            self._scroll = y + self.STUB_H - view_h

    def paintEvent(self, _event) -> None:
        painter = QPainter(self)
        painter.fillRect(self.rect(), Theme.paper_deep)
        painter.setPen(QPen(Theme.rail_line, 1))
        painter.drawLine(self.width() - 1, 0, self.width() - 1, self.height())

        mono = pick_font(FONT_MONO, "Consolas", 10)
        painter.setFont(mono)
        metrics = QFontMetrics(mono)

        for i, pn in enumerate(self._order):
            y = 4 + i * self.STUB_H - self._scroll
            if y + self.STUB_H < 0 or y > self.height():
                continue
            status = self._statuses.get(pn, "unreviewed")
            selected = pn == self._selected
            box = QRect(4, y, self.RAIL_W - 10, self.STUB_H - 4)
            if selected:
                painter.fillRect(box, Theme.surface)
                painter.setPen(QPen(Theme.accent, 2))
                painter.drawRect(box.adjusted(0, 0, -1, -1))
            else:
                painter.fillRect(box, Theme.paper)

            color = status_color(status)
            painter.setBrush(color if status != "unreviewed" else Qt.BrushStyle.NoBrush)
            painter.setPen(QPen(color, 2))
            painter.drawEllipse(box.left() + 8, box.center().y() - 5, 10, 10)

            text_left = box.left() + 26
            text_w = max(20, box.width() - 32)
            stub = metrics.elidedText(pn, Qt.TextElideMode.ElideRight, text_w)
            painter.setPen(Theme.ink)
            painter.drawText(
                text_left,
                box.top(),
                text_w,
                box.height(),
                Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft,
                stub,
            )

    def wheelEvent(self, event) -> None:
        delta = -event.angleDelta().y() // 4
        max_scroll = max(0, len(self._order) * self.STUB_H - self.height() + 8)
        self._scroll = max(0, min(max_scroll, self._scroll + delta))
        self.update()

    def mousePressEvent(self, event: QMouseEvent) -> None:
        if event.button() != Qt.MouseButton.LeftButton:
            return
        y = event.position().y() + self._scroll - 4
        idx = int(y // self.STUB_H)
        if 0 <= idx < len(self._order):
            self._window.focus_product(self._order[idx])
            self._window._refocus_window()


class ScoreMeter(QWidget):
    def __init__(self) -> None:
        super().__init__()
        self._score = 0.0
        self.setFixedHeight(10)

    def set_score(self, score: float) -> None:
        self._score = max(0.0, min(1.0, score))
        self.update()

    def paintEvent(self, _event) -> None:
        painter = QPainter(self)
        track = self.rect().adjusted(0, 2, 0, -2)
        painter.fillRect(track, Theme.paper_deep)
        fill_w = int(track.width() * self._score)
        if fill_w > 0:
            painter.fillRect(QRect(track.left(), track.top(), fill_w, track.height()), Theme.accent)


class TokenChipFlow(QWidget):
    """Flow of token chips for the merged candidate stream + missing-from-ref row."""

    def __init__(self) -> None:
        super().__init__()
        self._chips: list[tuple[str, str]] = []  # (label, kind) kind in shared|added|removed
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.setFixedHeight(24)

    def set_chips(self, chips: list[tuple[str, str]]) -> None:
        self._chips = chips
        self._apply_height()
        self.update()

    def resizeEvent(self, event) -> None:
        super().resizeEvent(event)
        self._apply_height()
        self.update()

    def relayout(self) -> None:
        self._apply_height()
        self.update()

    def _apply_height(self) -> None:
        width = max(self.width(), 1)
        _, total_h = self._layout_chips(width)
        if self.height() != total_h:
            self.setFixedHeight(total_h)

    def _layout_chips(self, width: int) -> tuple[list[tuple[QRect, str, str]], int]:
        font = pick_font(FONT_MONO, "Consolas", 13, QFont.Weight.DemiBold)
        metrics = QFontMetrics(font)
        pad_x, pad_y, gap, line_gap = 10, 7, 6, 8
        x, y = 0, 0
        row_h = metrics.height() + pad_y * 2
        placed: list[tuple[QRect, str, str]] = []
        usable = max(width, 40)
        for label, kind in self._chips:
            tw = metrics.horizontalAdvance(label) + pad_x * 2
            if x > 0 and x + tw > usable:
                x = 0
                y += row_h + line_gap
            placed.append((QRect(x, y, tw, row_h), label, kind))
            x += tw + gap
        total_h = y + row_h if placed else 24
        return placed, total_h

    def sizeHint(self) -> QSize:
        width = max(self.width(), 400)
        _, h = self._layout_chips(width)
        return QSize(width, h)

    def minimumSizeHint(self) -> QSize:
        return QSize(40, 24)

    def paintEvent(self, _event) -> None:
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        font = pick_font(FONT_MONO, "Consolas", 13, QFont.Weight.DemiBold)
        painter.setFont(font)
        placed, _ = self._layout_chips(max(self.width(), 1))
        for rect, label, kind in placed:
            if kind == "drawing":
                bg = Theme.chip_drawing_bg
                fg = Theme.chip_drawing_fg
                border = Theme.chip_drawing_border
                strike = False
                border_w = 3
            elif kind == "shared":
                bg = Theme.chip_shared_bg
                fg = Theme.chip_shared_fg
                border = Theme.chip_shared_border
                strike = False
                border_w = 2
            elif kind == "added":
                bg = Theme.chip_added_bg
                fg = Theme.chip_added_fg
                border = Theme.chip_added_bg
                strike = False
                border_w = 2
            else:  # removed
                bg = Theme.chip_removed_bg
                fg = Theme.chip_removed_fg
                border = Theme.chip_removed_bg
                strike = True
                border_w = 2
            painter.setBrush(bg)
            painter.setPen(QPen(border, border_w))
            painter.drawRoundedRect(rect, 3, 3)
            painter.setPen(fg)
            painter.drawText(rect, Qt.AlignmentFlag.AlignCenter, label)
            if strike:
                painter.setPen(QPen(fg, 2))
                mid = rect.center().y()
                painter.drawLine(rect.left() + 6, mid, rect.right() - 6, mid)


class MergedDiffHero(QWidget):
    """Large one-match compare: headers, score meter, candidate chips, missing chips."""

    def __init__(self) -> None:
        super().__init__()
        outer = QVBoxLayout(self)
        outer.setContentsMargins(24, 16, 28, 12)
        outer.setSpacing(6)

        self.ref_caption = QLabel("REFERENCE")
        self.ref_caption.setFont(pick_font(FONT_TITLE, "Segoe UI", 11, QFont.Weight.Bold))
        outer.addWidget(self.ref_caption)

        self.ref_pn = QLabel("")
        self.ref_pn.setFont(pick_font(FONT_MONO, "Consolas", 26, QFont.Weight.Bold))
        self.ref_pn.setWordWrap(True)
        outer.addWidget(self.ref_pn)

        self.ref_desc = QLabel("")
        self.ref_desc.setFont(pick_font(FONT_TITLE, "Segoe UI", 15))
        self.ref_desc.setWordWrap(True)
        self.ref_desc.setMinimumHeight(44)
        outer.addWidget(self.ref_desc)

        # Matched items (candidate tokens) sit directly under the reference
        # they were matched with; the candidate identity/score follows below.
        self.legend = QLabel(
            "shared  ·  drawing match xx-xx-xx (yellow)  ·  added (teal)  ·  missing (red)",
        )
        self.legend.setFont(pick_font(FONT_TITLE, "Segoe UI", 9))
        outer.addWidget(self.legend)

        self.stream_label = QLabel("CANDIDATE TOKENS")
        self.stream_label.setFont(pick_font(FONT_TITLE, "Segoe UI", 8, QFont.Weight.Bold))
        outer.addWidget(self.stream_label)

        self.cand_flow = TokenChipFlow()
        outer.addWidget(self.cand_flow)

        self.miss_label = QLabel("ONLY IN REFERENCE")
        self.miss_label.setFont(pick_font(FONT_TITLE, "Segoe UI", 8, QFont.Weight.Bold))
        outer.addWidget(self.miss_label)

        self.miss_flow = TokenChipFlow()
        outer.addWidget(self.miss_flow)

        self.rule = QFrame()
        self.rule.setFixedHeight(3)
        outer.addWidget(self.rule)

        self.cand_caption = QLabel("CANDIDATE")
        self.cand_caption.setFont(pick_font(FONT_TITLE, "Segoe UI", 10, QFont.Weight.Bold))
        outer.addWidget(self.cand_caption)

        head = QHBoxLayout()
        self.cand_pn = QLabel("")
        self.cand_pn.setFont(pick_font(FONT_MONO, "Consolas", 22, QFont.Weight.Bold))
        self.cand_pn.setWordWrap(True)
        head.addWidget(self.cand_pn, 1)

        self.status_chip = QLabel("")
        self.status_chip.setFont(pick_font(FONT_TITLE, "Segoe UI", 10, QFont.Weight.Bold))
        self.status_chip.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.status_chip.setFixedHeight(28)
        self.status_chip.setMinimumWidth(100)
        head.addWidget(self.status_chip, 0, Qt.AlignmentFlag.AlignTop)
        outer.addLayout(head)

        score_row = QHBoxLayout()
        self.score_label = QLabel("Score")
        self.score_label.setFont(pick_font(FONT_MONO, "Consolas", 10))
        score_row.addWidget(self.score_label)
        self.meter = ScoreMeter()
        score_row.addWidget(self.meter, 1)
        outer.addLayout(score_row)

        outer.addStretch(1)
        self.apply_theme()

    def apply_theme(self) -> None:
        self.ref_caption.setStyleSheet(
            f"color: {Theme.ink_muted.name()}; letter-spacing: 1px;",
        )
        self.ref_pn.setStyleSheet(f"color: {Theme.ink.name()};")
        self.ref_desc.setStyleSheet(
            f"color: {Theme.ink.name()}; background: {Theme.surface.name()}; "
            f"padding: 12px; border: 1px solid {Theme.paper_deep.name()};",
        )
        self.rule.setStyleSheet(f"background: {Theme.accent.name()}; border: none;")
        self.cand_caption.setStyleSheet(
            f"color: {Theme.ink_muted.name()}; letter-spacing: 1px;",
        )
        self.cand_pn.setStyleSheet(f"color: {Theme.ink.name()};")
        self.score_label.setStyleSheet(f"color: {Theme.ink_muted.name()};")
        self.legend.setStyleSheet(f"color: {Theme.ink_muted.name()};")
        self.stream_label.setStyleSheet(f"color: {Theme.ink_muted.name()};")
        self.miss_label.setStyleSheet(f"color: {Theme.coral.name()};")
        self.cand_flow.update()
        self.miss_flow.update()
        self.meter.update()

    def resizeEvent(self, event) -> None:
        super().resizeEvent(event)
        self.cand_flow.relayout()
        self.miss_flow.relayout()

    def clear(self) -> None:
        self.ref_pn.setText("—")
        self.ref_desc.setText("No cluster loaded.")
        self.cand_pn.setText("—")
        self.score_label.setText("Score —")
        self.meter.set_score(0.0)
        self.status_chip.setText("")
        self.status_chip.setStyleSheet("background: transparent;")
        self.cand_flow.set_chips([])
        self.miss_flow.set_chips([])

    def show_match(self, ref: dict, cand: dict, status: str) -> None:
        shared, only_ref, only_cand = token_diff(ref["description"], cand["description"])
        self.ref_pn.setText(ref["product_number"])
        self.ref_desc.setText(ref["description"] or "(empty)")
        self.cand_pn.setText(cand["product_number"])

        score = score_value(cand["score_to_parent"])
        self.score_label.setText(f"Score {format_score(cand['score_to_parent'])}")
        self.meter.set_score(score)

        status_label = STATUS_LABELS.get(status, "Unreviewed") if status != "unreviewed" else "Unreviewed"
        color = status_color(status)
        self.status_chip.setText(status_label.upper())
        if status != "unreviewed":
            self.status_chip.setStyleSheet(
                f"color: #F7F8FA; background: {color.name()}; "
                f"padding: 4px 10px; border-radius: 2px;",
            )
        else:
            self.status_chip.setStyleSheet(
                f"color: {Theme.ink_muted.name()}; background: {Theme.paper_deep.name()}; "
                f"padding: 4px 10px; border-radius: 2px;",
            )

        cand_chips: list[tuple[str, str]] = []
        for raw in (cand["description"] or "").split():
            upper = raw.upper()
            if upper in shared:
                # Drawing numbers shared with parent get max-contrast treatment
                if is_drawing_number(raw) or is_drawing_number(upper):
                    cand_chips.append((raw, "drawing"))
                else:
                    cand_chips.append((raw, "shared"))
            elif upper in only_cand:
                cand_chips.append((raw, "added"))
            else:
                cand_chips.append((raw, "shared"))
        self.cand_flow.set_chips(cand_chips)

        miss_chips: list[tuple[str, str]] = []
        seen: set[str] = set()
        for raw in (ref["description"] or "").split():
            upper = raw.upper()
            if upper in only_ref and upper not in seen:
                miss_chips.append((raw, "removed"))
                seen.add(upper)
        self.miss_flow.set_chips(miss_chips)
        QTimer.singleShot(0, self._relayout_flows)

    def _relayout_flows(self) -> None:
        self.cand_flow.relayout()
        self.miss_flow.relayout()


class DecisionBar(QWidget):
    def __init__(self, window: "ReviewWindow") -> None:
        super().__init__()
        self._window = window
        self.setFixedHeight(88)
        self._layout = QHBoxLayout(self)
        self._layout.setContentsMargins(16, 12, 16, 12)
        self._layout.setSpacing(12)
        self._buttons: list[tuple[QPushButton, str]] = []
        specs = (
            ("<-  DUPLICATE", "Left", "coral", "duplicate"),
            ("UNIQUE  ->", "Right", "teal", "unique"),
            ("DISCARD", "Up", "ochre", "discard"),
            ("UNREVIEW", "Down", "muted", "unreviewed"),
        )
        for title, hint, color_key, status in specs:
            btn = QPushButton(f"{title}\n{hint}")
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setFocusPolicy(Qt.FocusPolicy.NoFocus)
            btn.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
            btn.clicked.connect(lambda _=False, s=status: self._on_decide(s))
            self._layout.addWidget(btn)
            self._buttons.append((btn, color_key))
        self.apply_theme()

    def _color_for(self, key: str) -> QColor:
        return {
            "coral": Theme.coral,
            "teal": Theme.teal,
            "ochre": Theme.ochre,
            "muted": Theme.ink_muted,
        }[key]

    def apply_theme(self) -> None:
        for btn, color_key in self._buttons:
            color = self._color_for(color_key)
            hover_fg = "#F7F8FA" if not Theme.dark else Theme.paper.name()
            btn.setStyleSheet(f"""
                QPushButton {{
                    background: {Theme.surface.name()};
                    color: {Theme.ink.name()};
                    border: 2px solid {color.name()};
                    border-radius: 2px;
                    font-family: {FONT_TITLE};
                    font-size: 13pt;
                    font-weight: 700;
                    padding: 8px;
                }}
                QPushButton:hover {{
                    background: {color.name()};
                    color: {hover_fg};
                }}
            """)
        self.update()

    def _on_decide(self, status: str) -> None:
        if status == "unreviewed":
            self._window._clear_focused()
        else:
            self._window._mark_focused(status)
        self._window._refocus_window()

    def paintEvent(self, event) -> None:
        painter = QPainter(self)
        painter.fillRect(self.rect(), Theme.surface)
        painter.setPen(QPen(Theme.paper_deep, 1))
        painter.drawLine(0, 0, self.width(), 0)
        super().paintEvent(event)


class RelatedPanel(QWidget):
    """Right-side list of cross-cluster semantic suggestions for the current cluster.

    Purely mouse-driven (all widgets NoFocus) so it never interferes with the
    arrow-key review shortcuts. Jump opens that product's cluster; Pull moves it
    into the current cluster after an explicit human click (never automatic).
    """

    PANEL_W = 320

    def __init__(self, window: "ReviewWindow") -> None:
        super().__init__()
        self._window = window
        self.setFixedWidth(self.PANEL_W)
        self.setFocusPolicy(Qt.FocusPolicy.NoFocus)

        outer = QVBoxLayout(self)
        outer.setContentsMargins(10, 12, 12, 12)
        outer.setSpacing(8)

        self.caption = QLabel("RELATED (SEMANTIC)")
        self.caption.setFont(pick_font(FONT_TITLE, "Segoe UI", 10, QFont.Weight.Bold))
        outer.addWidget(self.caption)

        self.hint = QLabel("")
        self.hint.setFont(pick_font(FONT_TITLE, "Segoe UI", 9))
        self.hint.setWordWrap(True)
        outer.addWidget(self.hint)

        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setFrameShape(QFrame.Shape.NoFrame)
        self.scroll.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self._list_host = QWidget()
        self._list_layout = QVBoxLayout(self._list_host)
        self._list_layout.setContentsMargins(0, 0, 0, 0)
        self._list_layout.setSpacing(6)
        self._list_layout.addStretch(1)
        self.scroll.setWidget(self._list_host)
        outer.addWidget(self.scroll, 1)

        self._cards: list[QWidget] = []
        self.apply_theme()

    def _clear(self) -> None:
        for card in self._cards:
            card.setParent(None)
            card.deleteLater()
        self._cards = []

    def set_items(self, items: list[dict]) -> None:
        self._clear()
        if not items:
            self.hint.setText("No related items for this parent.")
        else:
            self.hint.setText(
                f"{len(items)} possible related item(s) in other clusters — "
                "Jump to inspect, or Pull into this cluster."
            )
        insert_at = self._list_layout.count() - 1  # keep the trailing stretch last
        for it in items:
            pn = it.get("suggested_product", "")
            score = format_score(it.get("semantic_score"))
            desc = (it.get("suggested_description") or "").strip()
            short = desc if len(desc) <= 70 else desc[:69] + "…"

            card = QWidget()
            card.setFocusPolicy(Qt.FocusPolicy.NoFocus)
            card_layout = QVBoxLayout(card)
            card_layout.setContentsMargins(0, 0, 0, 0)
            card_layout.setSpacing(4)

            info = QLabel(f"{pn}   ·   {score}\n{short or '(no description)'}")
            info.setWordWrap(True)
            info.setToolTip(f"{pn}\n{desc}")
            info.setFocusPolicy(Qt.FocusPolicy.NoFocus)
            card_layout.addWidget(info)

            row = QHBoxLayout()
            row.setContentsMargins(0, 0, 0, 0)
            row.setSpacing(6)
            jump_btn = QPushButton("Jump")
            jump_btn.setCursor(Qt.CursorShape.PointingHandCursor)
            jump_btn.setFocusPolicy(Qt.FocusPolicy.NoFocus)
            jump_btn.setToolTip(f"Open cluster for {pn}")
            jump_btn.clicked.connect(lambda _=False, p=pn: self._window._search_parent(p))
            pull_btn = QPushButton("Pull in")
            pull_btn.setCursor(Qt.CursorShape.PointingHandCursor)
            pull_btn.setFocusPolicy(Qt.FocusPolicy.NoFocus)
            pull_btn.setToolTip(
                f"Move {pn} into this cluster and mark as duplicate (human action only)"
            )
            pull_btn.clicked.connect(
                lambda _=False, p=pn, s=it: self._window.pull_related_into_cluster(p, s)
            )
            row.addWidget(jump_btn, 1)
            row.addWidget(pull_btn, 1)
            card_layout.addLayout(row)

            self._list_layout.insertWidget(insert_at, card)
            self._cards.append(card)
            insert_at += 1
        self._style_cards()

    def _style_cards(self) -> None:
        card_css = f"""
            QWidget {{
                background: {Theme.surface.name()};
                color: {Theme.ink.name()};
                border: 1px solid {Theme.paper_deep.name()};
                border-left: 3px solid {Theme.teal.name()};
                border-radius: 2px;
            }}
        """
        label_css = f"""
            QLabel {{
                background: transparent;
                color: {Theme.ink.name()};
                border: none;
                padding: 8px 10px 0 10px;
                font-family: {FONT_MONO};
                font-size: 9pt;
            }}
        """
        jump_css = f"""
            QPushButton {{
                background: {Theme.paper.name()};
                color: {Theme.ink.name()};
                border: 1px solid {Theme.paper_deep.name()};
                border-radius: 2px;
                padding: 6px 8px;
                font-family: {FONT_TITLE};
                font-size: 9pt;
            }}
            QPushButton:hover {{
                border-color: {Theme.accent.name()};
                background: {Theme.hover_tint};
            }}
        """
        pull_css = f"""
            QPushButton {{
                background: {Theme.teal.name()};
                color: #FFFFFF;
                border: 1px solid {Theme.teal.name()};
                border-radius: 2px;
                padding: 6px 8px;
                font-family: {FONT_TITLE};
                font-size: 9pt;
            }}
            QPushButton:hover {{
                border-color: {Theme.accent.name()};
            }}
        """
        for card in self._cards:
            card.setStyleSheet(card_css)
            for child in card.findChildren(QLabel):
                child.setStyleSheet(label_css)
            buttons = card.findChildren(QPushButton)
            if len(buttons) >= 2:
                buttons[0].setStyleSheet(jump_css)
                buttons[1].setStyleSheet(pull_css)
            elif buttons:
                buttons[0].setStyleSheet(jump_css)

    def apply_theme(self) -> None:
        self.caption.setStyleSheet(f"color: {Theme.ink.name()}; letter-spacing: 1px;")
        self.hint.setStyleSheet(f"color: {Theme.ink_muted.name()};")
        self.scroll.setStyleSheet("QScrollArea { background: transparent; border: none; }")
        self._list_host.setStyleSheet("background: transparent;")
        self._style_cards()
        self.update()

    def paintEvent(self, _event) -> None:
        painter = QPainter(self)
        painter.fillRect(self.rect(), Theme.paper_deep)
        painter.setPen(QPen(Theme.rail_line, 1))
        painter.drawLine(0, 0, 0, self.height())


class ReviewWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("Similarity Review")
        self.results_path = DEFAULT_RESULTS
        self.cluster_order: list[int] = []
        self.clusters: dict[int, list[dict]] = {}
        self.by_product: dict[str, dict] = {}
        self.semantic_suggestions: dict[str, list[dict]] = {}
        self.progress = load_progress(self.results_path)
        self.cluster_index = 0
        self.selected_product = ""
        self._reference_pn = ""
        self._candidate_order: list[str] = []
        self._screen = "review"
        self._timer_cluster_id: int | None = None
        self._timer_started_at: float | None = None

        shell = HatchPaper()
        root = QVBoxLayout(shell)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        self.top_strip = TopStrip(self)
        root.addWidget(self.top_strip)

        self.tab_strip = TabStrip(self)
        root.addWidget(self.tab_strip)

        self.stack = QStackedWidget()
        root.addWidget(self.stack, 1)

        # --- Review page ---
        review_page = QWidget()
        review_page.setStyleSheet("background: transparent;")
        review_layout = QVBoxLayout(review_page)
        review_layout.setContentsMargins(0, 0, 0, 0)
        review_layout.setSpacing(0)

        body = QHBoxLayout()
        body.setContentsMargins(0, 0, 0, 0)
        body.setSpacing(0)

        self.queue_rail = QueueRail(self)
        body.addWidget(self.queue_rail)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setStyleSheet("QScrollArea { background: transparent; border: none; }")
        self.hero = MergedDiffHero()
        self.hero.setStyleSheet("background: transparent;")
        scroll.setWidget(self.hero)
        body.addWidget(scroll, 1)

        self.related_panel = RelatedPanel(self)
        body.addWidget(self.related_panel)

        body_wrap = QWidget()
        body_wrap.setLayout(body)
        body_wrap.setStyleSheet("background: transparent;")
        review_layout.addWidget(body_wrap, 1)

        self.decision_bar = DecisionBar(self)
        review_layout.addWidget(self.decision_bar)

        self.reports_screen = ReportsScreen(self)

        self.stack.addWidget(review_page)
        self.stack.addWidget(self.reports_screen)

        self.setCentralWidget(shell)
        self.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
        self._install_shortcuts()
        # Route review keys at the application level so arrows/Tab/Space work no
        # matter which child widget (e.g. the scroll area) currently holds focus.
        app = QApplication.instance()
        if app is not None:
            app.installEventFilter(self)
        self._ui_clock = QTimer(self)
        self._ui_clock.setInterval(1000)
        self._ui_clock.timeout.connect(self._on_parent_clock_tick)
        self._ui_clock.start()
        dark = bool(self.progress.get("dark_mode", False))
        apply_theme_palette(dark=dark)
        _sync_legacy_color_aliases()
        self._apply_theme()
        self._refresh_top()
        # Maximized (not fullscreen) so the window is obvious on first launch.
        self.showMaximized()
        self.raise_()
        self.activateWindow()
        self._refocus_window()

    def show_review_screen(self) -> None:
        self._screen = "review"
        self.stack.setCurrentIndex(0)
        self._resume_parent_timer()
        self._refresh_top()
        self.top_strip.apply_theme()
        self.tab_strip.apply_theme()
        self._refocus_window()

    def show_reports_screen(self) -> None:
        self._pause_parent_timer(persist=True)
        self._screen = "reports"
        self.reports_screen.refresh()
        self.stack.setCurrentIndex(1)
        self.top_strip.apply_theme()
        self.tab_strip.apply_theme()
        self._refocus_window()

    def closeEvent(self, event) -> None:  # noqa: N802
        self._pause_parent_timer(persist=True)
        super().closeEvent(event)

    def _install_shortcuts(self) -> None:
        def bind(key: str, slot) -> None:
            sc = QShortcut(QKeySequence(key), self)
            sc.setContext(Qt.ShortcutContext.WindowShortcut)
            sc.activated.connect(slot)

        # Review navigation/marks are handled in keyPressEvent for reliable Tab/Space.
        bind("O", self._pick_results)
        bind("I", self._pick_and_run_input)
        bind("S", self.show_reports_screen)
        bind("R", self.show_reports_screen)
        bind("T", self._toggle_dark_mode)
        bind("Ctrl+F", self._focus_search)
        bind("G", self._toggle_related_panel)
        bind("Escape", self._toggle_fullscreen)

    def _review_keys_active(self) -> bool:
        """Only capture review keys when the review screen is truly interactive."""
        if self._screen != "review":
            return False
        if not self.isActiveWindow():
            return False
        # Never steal keys from a modal dialog (file/progress/message boxes).
        if QApplication.activeModalWidget() is not None:
            return False
        # Let text fields (e.g. the parent-code search) receive typing/navigation.
        focus = QApplication.focusWidget()
        return not isinstance(focus, (QLineEdit, QTextEdit))

    def _handle_review_key(self, event) -> bool:
        """Apply an arrow/Tab/Space review control. Returns True if consumed."""
        key = event.key()
        ctrl = bool(event.modifiers() & Qt.KeyboardModifier.ControlModifier)

        if key == Qt.Key.Key_Left:
            self._mark_focused("duplicate")
            return True
        if key == Qt.Key.Key_Right:
            self._mark_focused("unique")
            return True
        if key == Qt.Key.Key_Up:
            self._mark_focused("discard")
            return True
        if key == Qt.Key.Key_Down:
            self._clear_focused()
            return True
        if key == Qt.Key.Key_Tab:
            self._move_focus(-1 if ctrl else 1)
            return True
        if key == Qt.Key.Key_Backtab:
            self._move_focus(-1)
            return True
        if key == Qt.Key.Key_Space:
            if ctrl:
                self._prev_cluster()
            else:
                self._finish_cluster_and_next()
            return True
        return False

    def eventFilter(self, obj, event) -> bool:  # noqa: N802
        """Catch review controls before focused children (e.g. scroll area) eat them."""
        if event.type() == QEvent.Type.KeyPress and self._review_keys_active():
            if self._handle_review_key(event):
                event.accept()
                return True
        return super().eventFilter(obj, event)

    def keyPressEvent(self, event) -> None:  # noqa: N802
        """Fallback review controls (arrows / Tab / Space)."""
        if self._screen == "review" and self._handle_review_key(event):
            event.accept()
            return
        super().keyPressEvent(event)

    def focusNextPrevChild(self, next: bool) -> bool:  # noqa: A003, N802
        """Tab / Shift+Tab → next / previous child in the review queue."""
        if self._screen == "review":
            self._move_focus(1 if next else -1)
            return True
        return super().focusNextPrevChild(next)

    def _persist(self) -> None:
        """Save progress JSON + companion decisions Excel; never alter original results."""
        save_progress(self.progress, self.results_path)
        if self.results_path is None:
            return
        if not self.by_product:
            return
        try:
            save_decisions_workbook(self.results_path, self.by_product, self.progress)
        except OSError:
            # Excel lock / permission — JSON is already safe; ignore companion write.
            pass

    def _on_parent_clock_tick(self) -> None:
        if self._screen == "review" and self._timer_started_at is not None:
            self._refresh_top()

    def _accumulated_parent_seconds(self, cid: int | None) -> float:
        if cid is None:
            return 0.0
        times = self.progress.setdefault("parent_times", {})
        key = str(cid)
        try:
            return float(times.get(key, 0) or 0)
        except (TypeError, ValueError):
            return 0.0

    def _live_parent_seconds(self, cid: int | None = None) -> float:
        if cid is None:
            cid = self._timer_cluster_id
        seconds = self._accumulated_parent_seconds(cid)
        if (
            cid is not None
            and self._timer_cluster_id == cid
            and self._timer_started_at is not None
        ):
            seconds += max(0.0, time.monotonic() - self._timer_started_at)
        return seconds

    def _pause_parent_timer(self, *, persist: bool = True) -> None:
        if self._timer_cluster_id is None or self._timer_started_at is None:
            self._timer_started_at = None
            return
        elapsed = max(0.0, time.monotonic() - self._timer_started_at)
        times = self.progress.setdefault("parent_times", {})
        key = str(self._timer_cluster_id)
        try:
            prior = float(times.get(key, 0) or 0)
        except (TypeError, ValueError):
            prior = 0.0
        times[key] = prior + elapsed
        self._timer_started_at = None
        self._timer_cluster_id = None
        if persist:
            self._persist()

    def _resume_parent_timer(self, cid: int | None = None) -> None:
        if cid is None:
            cid = self._current_cluster_id()
        if cid is None or self._screen != "review":
            return
        if self._timer_cluster_id == cid and self._timer_started_at is not None:
            return
        if self._timer_started_at is not None:
            self._pause_parent_timer(persist=True)
        self._timer_cluster_id = cid
        self._timer_started_at = time.monotonic()

    def _refocus_window(self) -> None:
        self.setFocus(Qt.FocusReason.OtherFocusReason)

    def _toggle_fullscreen(self) -> None:
        if self.isFullScreen():
            self.showNormal()
        else:
            self.showFullScreen()
        self._refocus_window()

    def _toggle_dark_mode(self) -> None:
        apply_theme_palette(dark=not Theme.dark)
        _sync_legacy_color_aliases()
        self.progress["dark_mode"] = Theme.dark
        self._persist()
        self._apply_theme()
        self._show_current()
        self._refocus_window()

    def _apply_theme(self) -> None:
        self.setStyleSheet(f"""
            QMainWindow {{
                background-color: {Theme.paper.name()};
                color: {Theme.ink.name()};
            }}
            QScrollBar:vertical {{
                background: {Theme.paper.name()};
                width: 10px;
                margin: 0;
            }}
            QScrollBar::handle:vertical {{
                background: {Theme.paper_deep.name()};
                min-height: 24px;
                border-radius: 2px;
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0;
            }}
        """)
        self.top_strip.apply_theme()
        self.tab_strip.apply_theme()
        self.decision_bar.apply_theme()
        self.hero.apply_theme()
        self.related_panel.apply_theme()
        self.reports_screen.apply_theme()
        self.queue_rail.update()
        self.centralWidget().update()
        app = QApplication.instance()
        if app is not None:
            palette = app.palette()
            palette.setColor(palette.ColorRole.Window, Theme.paper)
            palette.setColor(palette.ColorRole.WindowText, Theme.ink)
            palette.setColor(palette.ColorRole.Base, Theme.surface)
            palette.setColor(palette.ColorRole.Text, Theme.ink)
            palette.setColor(palette.ColorRole.Button, Theme.surface)
            palette.setColor(palette.ColorRole.ButtonText, Theme.ink)
            palette.setColor(palette.ColorRole.Highlight, Theme.accent)
            palette.setColor(palette.ColorRole.HighlightedText, Theme.ink)
            app.setPalette(palette)

    def _show_stats(self) -> None:
        self.show_reports_screen()

    def _pick_results(self) -> None:
        from PySide6.QtWidgets import QFileDialog

        start = str(self.results_path.parent if self.results_path else APP_DIR)
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Open similarity results workbook",
            start,
            "Excel files (*.xlsx);;All files (*.*)",
        )
        if path:
            self._load_results(Path(path))
        self._refocus_window()

    def _startup_require_input(self) -> None:
        """Force a product Excel choice before review begins."""
        self.raise_()
        self.activateWindow()
        QMessageBox.information(
            self,
            "Choose input file",
            "Select a product export Excel to analyze.\n\n"
            "Accepted columns (FOExport style):\n"
            "  • Product number\n"
            "  • Product name\n\n"
            "This can take a few minutes on ~30k rows.",
        )
        self._pick_and_run_input(required=True)

    def _open_path_smart(self, path: Path) -> None:
        """Open a results workbook, or treat the file as product input and run."""
        if not path.exists():
            QMessageBox.critical(self, "Missing file", f"Could not find:\n{path}")
            self._startup_require_input()
            return
        try:
            if workbook_has_sheet(path, "Grouped Review"):
                self._load_results(path)
                return
        except Exception:
            pass
        self._run_similarity_on(path, required=True)

    def _pick_and_run_input(self, required: bool = False) -> None:
        """Choose a product Excel file, run similarity.py on it, then open the results."""
        from PySide6.QtWidgets import QFileDialog

        INPUT_FOLDER.mkdir(parents=True, exist_ok=True)
        title = "Choose product input Excel (required)" if required else "Choose product input Excel"
        path, _ = QFileDialog.getOpenFileName(
            self,
            title,
            str(INPUT_FOLDER),
            "Excel files (*.xlsx);;All files (*.*)",
        )
        if not path:
            if required:
                QMessageBox.warning(
                    self,
                    "Input required",
                    "An input Excel file is required to start.\nThe app will now close.",
                )
                QApplication.instance().quit()
            else:
                self._refocus_window()
            return

        self._run_similarity_on(Path(path), required=required)

    def _run_similarity_on(self, in_path: Path, *, required: bool = False) -> None:
        from PySide6.QtWidgets import QProgressDialog
        import os
        import shutil
        import subprocess

        work_dir = runtime_work_dir()
        work_input_dir = work_dir / "input"
        work_input_dir.mkdir(parents=True, exist_ok=True)

        # Copy input into a writable local folder (Downloads/OneDrive often deny child exes).
        local_input = work_input_dir / in_path.name
        try:
            if in_path.resolve() != local_input.resolve():
                shutil.copy2(in_path, local_input)
            else:
                local_input = in_path
        except OSError as exc:
            QMessageBox.critical(
                self,
                "Similarity failed",
                f"Could not copy input file:\n{in_path}\n\n{exc}\n\n"
                + access_denied_help(),
            )
            if required:
                self._pick_and_run_input(required=True)
            else:
                self._refocus_window()
            return

        out_path = work_dir / f"similarity_results__{local_input.stem}.xlsx"
        engine: Path | None = None

        wait = QProgressDialog(
            f"Running similarity on:\n{in_path.name}\n\nPreparing…",
            None,
            0,
            0,
            self,
        )
        wait.setWindowTitle("Similarity run")
        wait.setWindowModality(Qt.WindowModality.ApplicationModal)
        wait.setMinimumDuration(0)
        wait.show()
        QApplication.processEvents()

        if getattr(sys, "frozen", False):
            try:
                def _progress(msg: str) -> None:
                    wait.setLabelText(f"Running similarity on:\n{in_path.name}\n\n{msg}")
                    QApplication.processEvents()

                engine = ensure_local_engine(progress_cb=_progress)
            except Exception as exc:
                wait.close()
                QMessageBox.critical(
                    self,
                    "Similarity failed",
                    f"{exc}\n\n{access_denied_help(packaged_engine_exe())}",
                )
                if required:
                    self._pick_and_run_input(required=True)
                else:
                    self._refocus_window()
                return

        cmd = [*similarity_command(engine), str(local_input), "-o", str(out_path), "--gui"]
        log_lines: list[str] = []
        returncode = 1
        try:
            env = os.environ.copy()
            env["PYTHONUNBUFFERED"] = "1"
            cwd = str(engine.parent) if engine is not None else str(APP_DIR)
            proc = subprocess.Popen(
                cmd,
                cwd=cwd,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                env=env,
                bufsize=1,
            )
            assert proc.stdout is not None
            for line in proc.stdout:
                text = line.rstrip()
                if text:
                    log_lines.append(text)
                    wait.setLabelText(
                        f"Running similarity on:\n{in_path.name}\n\n{text}"
                    )
                    QApplication.processEvents()
            proc.wait()
            returncode = proc.returncode or 0
        except PermissionError as exc:
            wait.close()
            QMessageBox.critical(
                self,
                "Similarity failed",
                f"{exc}\n\n{access_denied_help(engine)}",
            )
            if required:
                self._pick_and_run_input(required=True)
            else:
                self._refocus_window()
            return
        except OSError as exc:
            wait.close()
            winerr = getattr(exc, "winerror", None)
            detail = access_denied_help(engine) if winerr == 5 else str(exc)
            QMessageBox.critical(
                self,
                "Similarity failed",
                f"{exc}\n\n{detail}",
            )
            if required:
                self._pick_and_run_input(required=True)
            else:
                self._refocus_window()
            return
        except Exception as exc:
            wait.close()
            QMessageBox.critical(self, "Similarity failed", str(exc))
            if required:
                self._pick_and_run_input(required=True)
            else:
                self._refocus_window()
            return
        finally:
            wait.close()

        if returncode != 0:
            detail = "\n".join(log_lines[-40:]) or "Unknown error"
            extra = ""
            if "Access is denied" in detail or "WinError 5" in detail:
                extra = "\n\n" + access_denied_help(engine)
            QMessageBox.critical(
                self,
                "Similarity failed",
                f"Could not process:\n{in_path}\n\n{detail[-2000:]}{extra}",
            )
            if required:
                self._pick_and_run_input(required=True)
            else:
                self._refocus_window()
            return

        if not out_path.exists():
            QMessageBox.critical(
                self,
                "Similarity failed",
                f"Engine finished but output was not created:\n{out_path}",
            )
            if required:
                self._pick_and_run_input(required=True)
            else:
                self._refocus_window()
            return

        QMessageBox.information(
            self,
            "Similarity complete",
            f"Results written to:\n{out_path}\n\nLoading into the reviewer…",
        )
        self._load_results(out_path)
        self.show_review_screen()
        self._refocus_window()

    def _load_results(self, path: Path) -> None:
        if not path.exists():
            QMessageBox.critical(self, "Missing file", f"Could not find:\n{path}")
            return

        # Flush time onto the previous workbook before swapping data/progress.
        self._pause_parent_timer(persist=True)

        try:
            cluster_order, clusters, by_product = load_grouped_review(path)
        except Exception as exc:
            QMessageBox.critical(self, "Load failed", f"Could not read Grouped Review sheet:\n{exc}")
            return

        self.cluster_order = cluster_order
        self.clusters = clusters
        self.by_product = by_product
        try:
            self.semantic_suggestions = load_semantic_suggestions(path)
        except Exception:
            self.semantic_suggestions = {}
        self.results_path = path
        self.progress = load_progress(path)
        self.progress.setdefault("decisions", {})
        self.progress.setdefault("clusters_completed", [])
        self.progress.setdefault("parent_times", {})
        self.progress.setdefault("cluster_moves", {})
        self.progress["source_file"] = path.name
        self.cluster_order = apply_cluster_moves(
            self.cluster_order,
            self.clusters,
            self.by_product,
            self.progress.get("cluster_moves", {}),
        )
        self._persist()
        self.cluster_index = self._first_incomplete_cluster_index()
        self.selected_product = ""
        self.setWindowTitle(f"Similarity Review — {path.name}")
        self._render_cluster()
        self._refresh_top()
        if self._screen == "reports":
            self.reports_screen.refresh()

    def _first_incomplete_cluster_index(self) -> int:
        completed = set(self.progress.get("clusters_completed", []))
        for i, cid in enumerate(self.cluster_order):
            if cid not in completed:
                return i
        return 0

    def _current_cluster_id(self) -> int | None:
        if not self.cluster_order:
            return None
        return self.cluster_order[self.cluster_index]

    def _item_status(self, product_number: str) -> str:
        decision = self.progress["decisions"].get(product_number)
        if not decision:
            return "unreviewed"
        return normalize_status(decision["status"])

    def _refresh_top(self) -> None:
        cid = self._current_cluster_id()
        if cid is None:
            self.top_strip.set_info("No clusters loaded", "")
            return
        items = self.clusters[cid]
        reviewed = sum(1 for it in items if it["product_number"] in self.progress["decisions"])
        done = cid in self.progress.get("clusters_completed", [])
        status = "DONE" if done else "in progress"
        marked_cands = sum(1 for pn in self._candidate_order if pn in self.progress["decisions"])
        parent_time = format_duration(self._live_parent_seconds(cid))
        self.top_strip.set_info(
            f"{self.results_path.name}  ·  Cluster {self.cluster_index + 1} / {len(self.cluster_order)}  ·  id {cid}  ·  {status}",
            f"{reviewed}/{len(items)} marked  ·  queue {marked_cands}/{len(self._candidate_order)}  ·  "
            f"all {len(self.progress['decisions']):,}/{len(self.by_product):,}  ·  "
            f"parent {parent_time}",
        )

    def _render_cluster(self) -> None:
        cid = self._current_cluster_id()
        if self._timer_cluster_id is not None and self._timer_cluster_id != cid:
            self._pause_parent_timer(persist=True)
        if self._screen == "review" and cid is not None:
            self._resume_parent_timer(cid)

        self._candidate_order = []
        self._reference_pn = ""

        if cid is None:
            self.hero.clear()
            self.queue_rail.set_queue([], {}, "")
            self._refresh_related()
            self._refresh_top()
            return

        items = self.clusters[cid]
        root = cluster_root(items)
        if root is None:
            self.hero.clear()
            self.queue_rail.set_queue([], {}, "")
            self._refresh_related()
            self._refresh_top()
            return

        self._reference_pn = root["product_number"]
        candidates = [i for i in items if i["product_number"] != self._reference_pn]
        candidates.sort(key=_score_sort_key)
        self._candidate_order = [c["product_number"] for c in candidates]

        statuses = {pn: self._item_status(pn) for pn in self._candidate_order}
        focus = self._pick_initial_focus()
        self.selected_product = focus
        self.queue_rail.set_queue(self._candidate_order, statuses, focus)
        self._show_current()
        self._refresh_related()
        self._refresh_top()
        self._refocus_window()

    def _refresh_related(self) -> None:
        """Populate the semantic panel with unique cross-cluster suggestions."""
        if not hasattr(self, "related_panel"):
            return
        cid = self._current_cluster_id()
        items: list[dict] = []
        if cid is not None and self.semantic_suggestions:
            best: dict[str, dict] = {}
            for member in self.clusters.get(cid, []):
                for sug in self.semantic_suggestions.get(member["product_number"], []):
                    sp = sug.get("suggested_product")
                    if not sp:
                        continue
                    # Skip if already in this cluster (including after a pull).
                    suggested_item = self.by_product.get(sp)
                    if suggested_item is not None and suggested_item.get("cluster_id") == cid:
                        continue
                    if sug.get("suggested_cluster_id") == cid:
                        continue
                    current = best.get(sp)
                    if current is None or score_value(sug.get("semantic_score")) > score_value(
                        current.get("semantic_score")
                    ):
                        best[sp] = sug
            items = sorted(
                best.values(),
                key=lambda s: score_value(s.get("semantic_score")),
                reverse=True,
            )[:SEMANTIC_MAX_PER_CLUSTER]
        self.related_panel.set_items(items)

    def pull_related_into_cluster(
        self,
        product_number: str,
        suggestion: dict | None = None,
    ) -> None:
        """Explicit human action: move a Related product into the current cluster."""
        if self._screen != "review":
            return
        cid = self._current_cluster_id()
        if cid is None:
            return
        item = self.by_product.get(product_number)
        if item is None:
            QMessageBox.warning(
                self,
                "Unknown product",
                f"Could not find product {product_number} in the loaded results.",
            )
            return
        if item.get("cluster_id") == cid:
            self._refresh_related()
            return

        ref = self._reference_pn or ""
        if not ref:
            root = cluster_root(self.clusters.get(cid, []))
            ref = root["product_number"] if root else ""

        old_cid = item.get("cluster_id")
        semantic_score = None
        if suggestion:
            semantic_score = suggestion.get("semantic_score")

        # Remove from old cluster
        if old_cid in self.clusters:
            self.clusters[old_cid] = [
                i for i in self.clusters[old_cid] if i["product_number"] != product_number
            ]
            if not self.clusters[old_cid]:
                del self.clusters[old_cid]
                if old_cid in self.cluster_order:
                    removed_at = self.cluster_order.index(old_cid)
                    self.cluster_order.pop(removed_at)
                    if removed_at < self.cluster_index:
                        self.cluster_index = max(0, self.cluster_index - 1)
                    elif self.cluster_index >= len(self.cluster_order):
                        self.cluster_index = max(0, len(self.cluster_order) - 1)

        item["cluster_id"] = cid
        if ref:
            item["linked_to_product"] = ref
        if item.get("depth", 0) == 0:
            item["depth"] = 1
        if semantic_score is not None and semantic_score != "":
            try:
                item["score_to_parent"] = float(semantic_score)
            except (TypeError, ValueError):
                pass

        members = self.clusters.setdefault(cid, [])
        if not any(i["product_number"] == product_number for i in members):
            members.append(item)
        _refresh_cluster_sizes(self.clusters)

        moves = self.progress.setdefault("cluster_moves", {})
        moves[product_number] = {
            "cluster_id": cid,
            "from_cluster_id": old_cid,
            "linked_to_product": ref,
            "semantic_score": semantic_score,
            "updated_at": utc_now(),
        }
        self.progress["decisions"][product_number] = {
            "status": "duplicate",
            "note": "pulled_from_related",
            "cluster_id": cid,
            "updated_at": utc_now(),
        }
        # Keep current cluster index pointing at the same id after order changes
        if cid in self.cluster_order:
            self.cluster_index = self.cluster_order.index(cid)

        self._persist()
        self.selected_product = product_number
        self._render_cluster()
        self._refocus_window()

    def _toggle_related_panel(self) -> None:
        if not hasattr(self, "related_panel"):
            return
        self.related_panel.setVisible(not self.related_panel.isVisible())
        self._refocus_window()

    def _pick_initial_focus(self) -> str:
        if not self._candidate_order:
            return ""
        if self.selected_product and self.selected_product in self._candidate_order:
            return self.selected_product
        for pn in self._candidate_order:
            if pn not in self.progress["decisions"]:
                return pn
        return self._candidate_order[0]

    def focus_product(self, product_number: str) -> None:
        if product_number not in self._candidate_order:
            return
        self.selected_product = product_number
        self.queue_rail.set_selected(product_number)
        self._show_current()

    def _show_current(self) -> None:
        ref = self.by_product.get(self._reference_pn)
        cand = self.by_product.get(self.selected_product)
        if not ref or not cand:
            self.hero.clear()
            if ref and not self._candidate_order:
                self.hero.ref_pn.setText(ref["product_number"])
                self.hero.ref_desc.setText(ref["description"] or "(empty)")
                self.hero.cand_pn.setText("No candidates in this cluster")
            return
        self.hero.show_match(ref, cand, self._item_status(self.selected_product))

    def _mark_focused(self, status: str) -> None:
        if self._screen != "review":
            return
        if self.selected_product and self.selected_product in self._candidate_order:
            self.mark_product(self.selected_product, status)

    def _clear_focused(self) -> None:
        if self._screen != "review":
            return
        if self.selected_product and self.selected_product in self._candidate_order:
            self.clear_product(self.selected_product)

    def _move_focus(self, delta: int) -> None:
        if self._screen != "review":
            return
        if not self._candidate_order:
            return
        try:
            idx = self._candidate_order.index(self.selected_product)
        except ValueError:
            idx = 0
        idx = max(0, min(len(self._candidate_order) - 1, idx + delta))
        self.focus_product(self._candidate_order[idx])

    def mark_product(self, product_number: str, status: str) -> None:
        item = self.by_product.get(product_number)
        if item is None:
            return

        status = normalize_status(status)
        self.selected_product = product_number
        self.progress["decisions"][product_number] = {
            "status": status,
            "note": "",
            "cluster_id": item["cluster_id"],
            "updated_at": utc_now(),
        }
        self._persist()
        self.queue_rail.update_status(product_number, status)
        self._refresh_top()
        next_pn = self._next_unreviewed_after(product_number)
        if next_pn == product_number:
            self._show_current()
        else:
            self.focus_product(next_pn)
        self._refocus_window()

    def clear_product(self, product_number: str) -> None:
        """Remove a decision so the item is unreviewed again."""
        if product_number not in self.by_product:
            return
        if product_number in self.progress["decisions"]:
            del self.progress["decisions"][product_number]
            self._persist()
        self.selected_product = product_number
        self.queue_rail.update_status(product_number, "unreviewed")
        self.queue_rail.set_selected(product_number)
        self._show_current()
        self._refresh_top()
        self._refocus_window()

    def _next_unreviewed_after(self, after_pn: str) -> str:
        if not self._candidate_order:
            return after_pn
        try:
            start = self._candidate_order.index(after_pn)
        except ValueError:
            start = -1
        n = len(self._candidate_order)
        for offset in range(1, n + 1):
            pn = self._candidate_order[(start + offset) % n]
            if pn not in self.progress["decisions"]:
                return pn
        return after_pn

    def _advance_to_next_unreviewed(self, after_pn: str) -> None:
        self.focus_product(self._next_unreviewed_after(after_pn))

    def _mark_cluster_complete(self) -> None:
        cid = self._current_cluster_id()
        if cid is None:
            return
        completed = self.progress.setdefault("clusters_completed", [])
        if cid not in completed:
            completed.append(cid)
            self._persist()

    def _finish_cluster_and_next(self) -> None:
        if self._screen != "review":
            return
        if not self.cluster_order:
            return
        self._mark_cluster_complete()
        if self.cluster_index >= len(self.cluster_order) - 1:
            self._render_cluster()
            QMessageBox.information(
                self,
                "All clusters done",
                "That was the last cluster. Use Ctrl+Space or Prev parent to go back if needed.",
            )
            return
        self.cluster_index += 1
        self.selected_product = ""
        self._render_cluster()

    def _prev_cluster(self) -> None:
        if self._screen != "review":
            return
        if not self.cluster_order or self.cluster_index <= 0:
            return
        self.cluster_index -= 1
        self.selected_product = ""
        self._render_cluster()
        self._refocus_window()

    def _match_cluster(self, query: str) -> tuple[int | None, str | None]:
        """Find a cluster for a parent/product code.

        Returns (cluster_index, product_to_focus). product_to_focus is None when
        the match is the parent itself, or the candidate product number to select.
        """
        q = query.strip().upper()
        if not q or not self.cluster_order:
            return None, None

        roots: list[tuple[int, int, str]] = []
        for idx, cid in enumerate(self.cluster_order):
            root = cluster_root(self.clusters.get(cid, []))
            roots.append((idx, cid, root["product_number"] if root else ""))

        # 1) Exact parent (cluster reference) match.
        for idx, _cid, rpn in roots:
            if rpn.upper() == q:
                return idx, None

        # 2) Exact match on any product — jump to its cluster and focus it.
        for idx, cid, rpn in roots:
            for item in self.clusters.get(cid, []):
                pn = item["product_number"]
                if pn.upper() == q:
                    return idx, (None if pn == rpn else pn)

        # 3) Prefix match on a parent code.
        for idx, _cid, rpn in roots:
            if rpn.upper().startswith(q):
                return idx, None

        # 4) Substring match on any product.
        for idx, cid, rpn in roots:
            for item in self.clusters.get(cid, []):
                pn = item["product_number"]
                if q in pn.upper():
                    return idx, (None if pn == rpn else pn)

        return None, None

    def _search_parent(self, query: str) -> None:
        """Jump to the cluster whose parent (or member) matches the code."""
        text = (query or "").strip()
        if not text:
            self._refocus_window()
            return
        if not self.cluster_order:
            QMessageBox.information(self, "No data", "Load a results workbook first.")
            return

        idx, focus_pn = self._match_cluster(text)
        if idx is None:
            QMessageBox.information(
                self,
                "Not found",
                f"No parent or product matches:\n{text}",
            )
            self._refocus_window()
            return

        if self._screen != "review":
            self.show_review_screen()
        self.cluster_index = idx
        self.selected_product = focus_pn or ""
        self._render_cluster()
        if focus_pn:
            self.focus_product(focus_pn)
        self._refocus_window()

    def _focus_search(self) -> None:
        if self._screen != "review":
            self.show_review_screen()
        self.tab_strip.focus_search()


def _crash_log_path() -> Path:
    return APP_DIR / "review_crash.log"


def _write_crash_log(text: str) -> Path:
    path = _crash_log_path()
    try:
        path.write_text(text, encoding="utf-8")
    except OSError:
        path = Path.cwd() / "review_crash.log"
        path.write_text(text, encoding="utf-8")
    return path


def main() -> None:
    _boot_log("main() start")
    app = QApplication(sys.argv)
    _boot_log("QApplication created")
    app.setStyle("Fusion")
    apply_theme_palette(dark=False)
    _sync_legacy_color_aliases()
    palette = app.palette()
    palette.setColor(palette.ColorRole.Window, Theme.paper)
    palette.setColor(palette.ColorRole.WindowText, Theme.ink)
    palette.setColor(palette.ColorRole.Base, Theme.surface)
    palette.setColor(palette.ColorRole.Text, Theme.ink)
    palette.setColor(palette.ColorRole.Button, Theme.surface)
    palette.setColor(palette.ColorRole.ButtonText, Theme.ink)
    palette.setColor(palette.ColorRole.Highlight, Theme.accent)
    palette.setColor(palette.ColorRole.HighlightedText, Theme.ink)
    app.setPalette(palette)
    window = ReviewWindow()
    _boot_log("ReviewWindow created")
    # Optional CLI: results workbook OR product input Excel (e.g. FOExport.xlsx)
    if len(sys.argv) > 1:
        candidate = Path(sys.argv[1])
        if candidate.exists():
            QTimer.singleShot(0, lambda: window._open_path_smart(candidate))
        else:
            QTimer.singleShot(0, window._startup_require_input)
    else:
        QTimer.singleShot(0, window._startup_require_input)
    window.show()
    window.raise_()
    window.activateWindow()
    _boot_log("entering event loop")
    sys.exit(app.exec())


if __name__ == "__main__":
    try:
        main()
    except Exception:
        import traceback

        detail = traceback.format_exc()
        log_path = _write_crash_log(detail)
        try:
            # May fail if Qt itself never started.
            from PySide6.QtWidgets import QApplication, QMessageBox

            app = QApplication.instance() or QApplication(sys.argv)
            QMessageBox.critical(
                None,
                "Similarity Review failed to start",
                f"Could not start the app.\n\nDetails saved to:\n{log_path}",
            )
        except Exception:
            print(detail, file=sys.stderr)
            print(f"Crash log: {log_path}", file=sys.stderr)
        sys.exit(1)
