"""
Similarity check for ~30k descriptions.

Method
------
  1. Normalize each description: UPPERCASE, split on whitespace, dedupe -> token set.
     The set is then joined back as a single alphabetically-sorted string
     ("canonical form") so the same words always show up in the same order
     and rows can be eyeballed side-by-side.
  2. Compute exact Jaccard similarity between every pair of rows:
         J(A, B) = |A ∩ B| / |A ∪ B|
  3. Speed tricks:
       - inverted index so only pairs that share a token are scored
       - drop ultra-common tokens (MAX_DOC_FREQUENCY)
       - collapse identical token sets and score each unique text once
       - score each unique pair from one side only (j > i)
  4. Union-find on every pair >= SIMILARITY_THRESHOLD gives clusters of
     near-identical descriptions.
  5. Rows whose canonical form is byte-identical are also reported as a
     dedicated "Exact Duplicates" sheet — that's the strictest form of
     match (same bag of words, regardless of original word order).

Output: one .xlsx file — start on the "Summary" sheet, then work through:
  - "Summary"          : counts and where to begin manual review
  - "Grouped Review"   : each product once; similar items nested underneath
  - "Exact Duplicates" : strict duplicate groups (same bag of words)
  - "Review Pairs"     : one row per pair, descriptions side-by-side + token diff
  - "Needs Review"     : flat list of cluster members (legacy layout)
  - "Clusters"         : all rows grouped by cluster, colour-coded

Optional legacy sheets (off by default):
  - "Top Matches"      : wide layout, top-N neighbours per row
  - "All Pairs"        : every pair >= threshold

Edit the CONFIG block below, then run:
    pip install -r requirements.txt
    python similarity.py
    python similarity.py path\\to\\products.xlsx
    python similarity.py path\\to\\products.xlsx -o my_results.xlsx
"""

from __future__ import annotations

import sys
import time
from bisect import bisect_right
from collections import Counter, defaultdict
from pathlib import Path

import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ===== CONFIG ================================================================

# Drop your .xlsx file into the "input" folder next to this script.
# If there are multiple .xlsx files in there, the most recently modified one
# is used. To force a specific file, set INPUT_FILE to its name (it can be
# relative to the input folder, e.g. "Products_xyz.xlsx", or an absolute path).
INPUT_FOLDER = "input"
INPUT_FILE: str | None = None               # None = auto-pick newest from INPUT_FOLDER

SHEET_NAME: str | int = 0                   # sheet index (0 = first) or sheet name
DESCRIPTION_COLUMN: str | None = None       # column name; None = auto-detect
PRODUCT_NUMBER_COLUMN: str | None = None    # column name; None = auto-detect

OUTPUT_FILE = "similarity_results.xlsx"

SIMILARITY_THRESHOLD = 0.50   # 0..1; pairs >= this are "matches" and join a cluster
TOP_N_MATCHES = 10            # how many neighbours to keep per row (legacy wide sheet)
REVIEW_PAIRS_MIN_SCORE = 0.50 # minimum score on the readable "Review Pairs" sheet
EMIT_WIDE_TOP_MATCHES = False # legacy 67-column sheet — usually not needed
EMIT_ALL_PAIRS_SHEET = False  # full pair dump — use Review Pairs instead

# Optional: ignore tokens that appear in more than this fraction of all rows.
# Common words like "THE", "AND" can blow up the inner loop without changing
# Jaccard scores much. 0.0 = keep every token. If the run is too slow, try 0.5.
MAX_DOC_FREQUENCY = 0.0

# Semantic "possible related" suggestions ------------------------------------
# A sentence-transformers model embeds every description and finds near
# neighbours that landed in a DIFFERENT cluster. These never change the strict
# Jaccard clusters; they are written to a "Semantic Suggestions" sheet and shown
# in the reviewer as click-to-jump hints so a human can catch vague synonym
# cases (e.g. "ZIP TIE" vs "CABLE TIE") the lexical matcher misses.
SEMANTIC_ENABLED = True
# Path to a bundled model folder (offline). Absolute, or relative to this file
# / the PyInstaller bundle. Pre-download all-MiniLM-L6-v2 into models/.
SEMANTIC_MODEL_PATH = "models/all-MiniLM-L6-v2"
SEMANTIC_THRESHOLD = 0.55   # cosine 0..1; higher = stricter / fewer suggestions
SEMANTIC_TOP_K = 5          # max cross-cluster suggestions kept per product
SEMANTIC_BATCH_SIZE = 256   # embedding batch size

# =============================================================================


DESCRIPTION_NAME_HINTS = (
    "product name", "search name", "description", "desc", "text",
    "name", "title", "item",
)
PRODUCT_NAME_HINTS = (
    "product number", "product_number", "product no", "product_no", "product #",
    "productno", "productnumber", "product",
    "part number", "part_number", "part no", "part_no", "partno", "part",
    "item number", "item_number", "item no", "item_no", "itemno",
    "pn", "p/n", "p_n", "sku", "code", "id",
)

# Known Dynamics / FO export layout (as used by FOExport.xlsx).
FO_EXPORT_PRODUCT_COL = "Product number"
FO_EXPORT_NAME_COL = "Product name"
FO_EXPORT_SEARCH_COL = "Search name"


def _norm_header(value: object) -> str:
    return str(value).strip().lower()


def detect_fo_export(df: pd.DataFrame) -> bool:
    """True when workbook looks like FOExport (Product number + Product name)."""
    cols = {_norm_header(c) for c in df.columns}
    return "product number" in cols and "product name" in cols


def resolve_columns(df: pd.DataFrame) -> tuple[str, str, str | None]:
    """Return (product_col, description_col, optional_search_col)."""
    by_low = {_norm_header(c): c for c in df.columns}
    if detect_fo_export(df):
        return (
            by_low["product number"],
            by_low["product name"],
            by_low.get("search name"),
        )

    prod_col = PRODUCT_NUMBER_COLUMN or autodetect_product_column(df)
    if prod_col is None:
        raise ValueError(
            "could not auto-detect a product-number column.\n"
            f"  Available columns: {list(df.columns)}\n"
            "  Expected FOExport columns: 'Product number', 'Product name'."
        )
    desc_col = DESCRIPTION_COLUMN or autodetect_description_column(df, exclude=prod_col)
    if desc_col == prod_col:
        raise ValueError(
            "description column and product column resolved to the same column.\n"
            f"  Available columns: {list(df.columns)}"
        )
    search_col = by_low.get("search name")
    return prod_col, desc_col, search_col


def clean_product_id(value: object) -> str:
    """Stable product id string (avoid 1010000010.0 from Excel floats)."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    text = str(value).strip()
    if text.endswith(".0") and text.replace(".", "", 1).isdigit():
        return text[:-2]
    return text


def build_descriptions(
    df: pd.DataFrame,
    name_col: str,
    search_col: str | None,
) -> list[str]:
    """Product name, falling back to Search name when name is blank."""
    names = df[name_col]
    searches = df[search_col] if search_col and search_col in df.columns else None
    out: list[str] = []
    for i in range(len(df)):
        raw = names.iloc[i]
        text = "" if pd.isna(raw) else str(raw).strip()
        if not text and searches is not None:
            alt = searches.iloc[i]
            text = "" if pd.isna(alt) else str(alt).strip()
        out.append(text)
    return out


def _columns_lower(df: pd.DataFrame) -> list[tuple[str, str]]:
    """List of (lowercased_name, original_name), order preserved."""
    return [(str(c).strip().lower(), c) for c in df.columns]


def _match_hint(
    cols: list[tuple[str, str]],
    hints: tuple[str, ...],
    exclude: str | None = None,
) -> str | None:
    """Find a column whose name matches one of the hints.
    1) Exact match on the whole column name (in hint order).
    2) Substring match — longest hints first, so 'product number'
       wins over the generic 'product'.
    """
    for h in hints:
        for low, orig in cols:
            if orig == exclude:
                continue
            if low == h:
                return orig
    for h in sorted(hints, key=len, reverse=True):
        for low, orig in cols:
            if orig == exclude:
                continue
            if h in low:
                return orig
    return None


def autodetect_product_column(df: pd.DataFrame) -> str | None:
    return _match_hint(_columns_lower(df), PRODUCT_NAME_HINTS)


def autodetect_description_column(df: pd.DataFrame, exclude: str | None) -> str:
    cols = _columns_lower(df)
    hit = _match_hint(cols, DESCRIPTION_NAME_HINTS, exclude=exclude)
    if hit is not None:
        return hit
    for col in df.columns:
        if col == exclude:
            continue
        if df[col].dtype == object:
            return col
    return df.columns[0]


def tokenize(text: str) -> tuple[frozenset[str], str]:
    """UPPERCASE, split on whitespace, dedupe. Returns the token set AND the
    canonical alphabetical form for eyeball comparison."""
    if not isinstance(text, str):
        return frozenset(), ""
    tokens = frozenset(text.upper().split())
    canonical = " ".join(sorted(tokens))
    return tokens, canonical


GROUP_FILLS = (
    PatternFill(start_color="E8F4FC", end_color="E8F4FC", fill_type="solid"),
    PatternFill(start_color="FFF4E5", end_color="FFF4E5", fill_type="solid"),
    PatternFill(start_color="EAF7EA", end_color="EAF7EA", fill_type="solid"),
    PatternFill(start_color="F3E8FD", end_color="F3E8FD", fill_type="solid"),
    PatternFill(start_color="FCE8E8", end_color="FCE8E8", fill_type="solid"),
    PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"),
)

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")

EXACT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
HIGH_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")


def format_token_diff(ts_a: frozenset[str], ts_b: frozenset[str]) -> str:
    """Human-readable token breakdown for side-by-side review."""
    only_a = sorted(ts_a - ts_b)
    only_b = sorted(ts_b - ts_a)
    shared = sorted(ts_a & ts_b)
    parts: list[str] = []
    if only_a:
        parts.append("ONLY A: " + " ".join(only_a))
    if only_b:
        parts.append("ONLY B: " + " ".join(only_b))
    if shared:
        parts.append("SHARED: " + " ".join(shared))
    return " | ".join(parts)


def build_review_pairs_df(
    pair_edges: list[tuple[int, int, float]],
    products: list[str],
    descriptions: list[str],
    token_sets: list[frozenset[str]],
    min_score: float,
) -> pd.DataFrame:
    rows: list[dict] = []
    for i, j, score in pair_edges:
        if score < min_score:
            continue
        rows.append({
            "score": round(score, 4),
            "exact": score >= 0.9999,
            "product_a": products[i],
            "description_a": descriptions[i],
            "product_b": products[j],
            "description_b": descriptions[j],
            "token_diff": format_token_diff(token_sets[i], token_sets[j]),
            "row_a": i,
            "row_b": j,
        })
    if not rows:
        return pd.DataFrame(columns=[
            "score", "exact", "product_a", "description_a",
            "product_b", "description_b", "token_diff", "row_a", "row_b",
        ])
    out = pd.DataFrame(rows)
    return out.sort_values(["exact", "score"], ascending=[False, False]).reset_index(drop=True)


def build_adjacency_scores(
    pair_edges: list[tuple[int, int, float]],
) -> dict[int, dict[int, float]]:
    adj: dict[int, dict[int, float]] = defaultdict(dict)
    for i, j, score in pair_edges:
        adj[i][j] = score
        adj[j][i] = score
    return adj


def order_cluster_bfs(
    members: list[int],
    adj: dict[int, dict[int, float]],
) -> list[tuple[int, int, int | None, float | None]]:
    """BFS from one cluster root — depth-1 nodes are direct matches of the parent."""
    member_set = set(members)
    if len(members) == 1:
        return [(members[0], 0, None, None)]

    def degree(node: int) -> int:
        return sum(1 for nb in adj.get(node, {}) if nb in member_set)

    root = min(members, key=lambda n: (-degree(n), n))
    parent_of: dict[int, tuple[int | None, float | None]] = {root: (None, None)}
    depth_of: dict[int, int] = {root: 0}
    queue = [root]

    while queue:
        node = queue.pop(0)
        neighbors = [
            (nb, adj[node][nb])
            for nb in adj.get(node, {})
            if nb in member_set and nb not in parent_of
        ]
        neighbors.sort(key=lambda x: (-x[1], x[0]))
        for nb, sc in neighbors:
            parent_of[nb] = (node, sc)
            depth_of[nb] = depth_of[node] + 1
            queue.append(nb)

    for m in sorted(members):
        if m not in parent_of:
            parent_of[m] = (None, None)
            depth_of[m] = 0

    order: list[tuple[int, int, int | None, float | None]] = []
    seen: set[int] = set()
    bfs_queue = [root]
    while bfs_queue:
        node = bfs_queue.pop(0)
        if node in seen:
            continue
        seen.add(node)
        p, sc = parent_of[node]
        order.append((node, depth_of[node], p, sc))
        children = sorted(
            [nb for nb in adj.get(node, {}) if nb in member_set and nb not in seen],
            key=lambda nb: (-adj[node][nb], nb),
        )
        bfs_queue.extend(children)

    for m in sorted(members):
        if m not in seen:
            p, sc = parent_of[m]
            order.append((m, depth_of[m], p, sc))
    return order


def build_grouped_review_df(
    products: list[str],
    descriptions: list[str],
    cluster_ids: list[int],
    cluster_sizes: dict[int, int],
    pair_edges: list[tuple[int, int, float]],
    dup_group_by_row: dict[int, int],
    *,
    include_singletons: bool = False,
) -> pd.DataFrame:
    """One row per product — one cluster root, matches branch by BFS depth."""
    adj = build_adjacency_scores(pair_edges)
    by_cluster: dict[int, list[int]] = defaultdict(list)
    for i, cid in enumerate(cluster_ids):
        by_cluster[cid].append(i)

    cluster_order = sorted(
        by_cluster.keys(),
        key=lambda cid: (
            cluster_sizes[cid] <= 1,
            -cluster_sizes[cid],
            cid,
        ),
    )

    columns = [
        "cluster_id", "cluster_size", "position_in_cluster", "depth",
        "product_number", "description", "linked_to_product",
        "score_to_parent", "n_similar_in_cluster", "exact_dup_group",
    ]
    rows: list[dict] = []
    for cid in cluster_order:
        csize = cluster_sizes[cid]
        if csize <= 1 and not include_singletons:
            continue
        members = by_cluster[cid]
        member_set = set(members)
        for pos, (idx, depth, parent, score) in enumerate(
            order_cluster_bfs(members, adj), start=1,
        ):
            n_direct = sum(1 for nb in adj.get(idx, {}) if nb in member_set)
            rows.append({
                "cluster_id": cid,
                "cluster_size": csize,
                "position_in_cluster": pos,
                "depth": depth,
                "product_number": products[idx],
                "description": descriptions[idx],
                "linked_to_product": products[parent] if parent is not None else "",
                "score_to_parent": round(score, 4) if score is not None else None,
                "n_similar_in_cluster": n_direct,
                "exact_dup_group": dup_group_by_row.get(idx),
            })

    if not rows:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame(rows)[columns]


SEMANTIC_COLUMNS = [
    "product_number", "product_cluster_id",
    "suggested_product", "suggested_cluster_id",
    "suggested_description", "semantic_score",
]


def _resolve_semantic_model_path() -> str | None:
    """Locate the bundled sentence-transformers model without hitting the network."""
    if not SEMANTIC_MODEL_PATH:
        return None
    configured = Path(SEMANTIC_MODEL_PATH)
    if configured.is_absolute():
        return str(configured) if configured.exists() else None
    bases = [
        Path(getattr(sys, "_MEIPASS", "")) if getattr(sys, "frozen", False) else None,
        Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else None,
        Path(__file__).resolve().parent,
    ]
    for base in bases:
        if base is None:
            continue
        candidate = base / configured
        if candidate.exists():
            return str(candidate)
    return None


def build_semantic_suggestions_df(
    products: list[str],
    descriptions: list[str],
    cluster_ids: list[int],
) -> pd.DataFrame:
    """Cross-cluster nearest-neighbour suggestions via sentence embeddings.

    Returns rows of (product, its cluster, suggested product, suggested cluster,
    suggested description, cosine score). Suggestions are review-only hints and
    never affect the strict Jaccard clusters. Any failure degrades to empty.
    """
    n = len(products)
    if n < 2:
        return pd.DataFrame(columns=SEMANTIC_COLUMNS)

    import os

    # Never reach out to Hugging Face at runtime — the model is bundled.
    os.environ.setdefault("HF_HUB_OFFLINE", "1")
    os.environ.setdefault("TRANSFORMERS_OFFLINE", "1")

    try:
        import numpy as np
        from sentence_transformers import SentenceTransformer
    except Exception as exc:  # noqa: BLE001 - optional dependency / build
        print(f"  Semantic pass skipped (import failed): {exc}")
        return pd.DataFrame(columns=SEMANTIC_COLUMNS)

    model_path = _resolve_semantic_model_path()
    if model_path is None:
        print(f"  Semantic pass skipped: model not found at {SEMANTIC_MODEL_PATH!r}")
        return pd.DataFrame(columns=SEMANTIC_COLUMNS)

    try:
        model = SentenceTransformer(model_path, device="cpu")
    except Exception as exc:  # noqa: BLE001
        print(f"  Semantic pass skipped (model load failed): {exc}")
        return pd.DataFrame(columns=SEMANTIC_COLUMNS)

    # Collapse identical (case-folded) descriptions so each is embedded once.
    uniq_index: dict[str, int] = {}
    uniq_texts: list[str] = []
    uniq_rows: list[list[int]] = []
    for i, desc in enumerate(descriptions):
        key = (desc or "").strip().lower()
        u = uniq_index.get(key)
        if u is None:
            u = len(uniq_texts)
            uniq_index[key] = u
            uniq_texts.append(desc or "")
            uniq_rows.append([])
        uniq_rows[u].append(i)
    m = len(uniq_texts)

    print(f"Semantic pass: embedding {m:,} unique descriptions (model on CPU) ...")
    try:
        emb = model.encode(
            uniq_texts,
            batch_size=int(SEMANTIC_BATCH_SIZE),
            convert_to_numpy=True,
            normalize_embeddings=True,
            show_progress_bar=False,
        ).astype(np.float32)
    except Exception as exc:  # noqa: BLE001
        print(f"  Semantic pass skipped (encode failed): {exc}")
        return pd.DataFrame(columns=SEMANTIC_COLUMNS)

    thr = float(SEMANTIC_THRESHOLD)
    k = int(SEMANTIC_TOP_K)
    # Pull a few extra neighbours before the cross-cluster filter trims them.
    kbuf = min(m - 1, max(k * 4, k + 8))
    block = 512
    rows: list[dict] = []

    for start in range(0, m, block):
        stop = min(start + block, m)
        sims = emb[start:stop] @ emb.T  # (b, m), values in [-1, 1]
        for bi in range(stop - start):
            ui = start + bi
            row = sims[bi]
            row[ui] = -1.0  # never suggest self
            part = np.argpartition(row, -kbuf)[-kbuf:]
            part = part[np.argsort(row[part])[::-1]]
            neighbours: list[tuple[int, float]] = []
            for uj in part:
                score = float(row[uj])
                if score < thr:
                    break
                neighbours.append((int(uj), score))
            if not neighbours:
                continue
            for i in uniq_rows[ui]:
                ci = cluster_ids[i]
                kept = 0
                for uj, score in neighbours:
                    rep = next(
                        (j for j in uniq_rows[uj] if cluster_ids[j] != ci),
                        None,
                    )
                    if rep is None:
                        continue
                    rows.append({
                        "product_number": products[i],
                        "product_cluster_id": ci,
                        "suggested_product": products[rep],
                        "suggested_cluster_id": cluster_ids[rep],
                        "suggested_description": descriptions[rep],
                        "semantic_score": round(score, 4),
                    })
                    kept += 1
                    if kept >= k:
                        break
        done = min(stop, m)
        print(f"  semantic neighbours {done:,} / {m:,} unique")

    if not rows:
        return pd.DataFrame(columns=SEMANTIC_COLUMNS)
    return pd.DataFrame(rows)[SEMANTIC_COLUMNS]


def build_needs_review_df(
    products: list[str],
    descriptions: list[str],
    cluster_ids: list[int],
    cluster_sizes: dict[int, int],
    dup_groups: list[tuple[str, list[int]]],
) -> pd.DataFrame:
    rows: list[dict] = []
    for gid, (_canon, ids) in enumerate(dup_groups):
        for i in ids:
            rows.append({
                "review_type": "exact_duplicate",
                "group_id": gid,
                "group_size": len(ids),
                "product_number": products[i],
                "description": descriptions[i],
                "cluster_id": cluster_ids[i],
                "cluster_size": cluster_sizes[cluster_ids[i]],
            })
    seen_dups = {i for _, ids in dup_groups for i in ids}
    for i, cid in enumerate(cluster_ids):
        if cluster_sizes[cid] <= 1:
            continue
        if i in seen_dups:
            continue
        rows.append({
            "review_type": "cluster_member",
            "group_id": cid,
            "group_size": cluster_sizes[cid],
            "product_number": products[i],
            "description": descriptions[i],
            "cluster_id": cid,
            "cluster_size": cluster_sizes[cid],
        })
    if not rows:
        return pd.DataFrame(columns=[
            "review_type", "group_id", "group_size", "product_number",
            "description", "cluster_id", "cluster_size",
        ])
    type_order = {"exact_duplicate": 0, "cluster_member": 1}
    out = pd.DataFrame(rows)
    out["_type_order"] = out["review_type"].map(type_order)
    out = out.sort_values(
        ["_type_order", "group_size", "group_id", "product_number"],
        ascending=[True, False, True, True],
    ).drop(columns="_type_order").reset_index(drop=True)
    return out


def build_summary_df(
    *,
    n: int,
    pair_edges: list[tuple[int, int, float]],
    review_pairs_df: pd.DataFrame,
    grouped_review_df: pd.DataFrame,
    dup_groups: list[tuple[str, list[int]]],
    cluster_sizes: dict[int, int],
    needs_review_df: pd.DataFrame,
    similarity_threshold: float,
    review_pairs_min_score: float,
    input_file: Path,
) -> pd.DataFrame:
    n_pairs = len(pair_edges)
    n_exact_pairs = sum(1 for _, _, s in pair_edges if s >= 0.9999)
    n_dup_groups = len(dup_groups)
    rows_in_dups = sum(len(ids) for _, ids in dup_groups)
    multi_clusters = sum(1 for s in cluster_sizes.values() if s > 1)
    rows_in_multi = sum(s for s in cluster_sizes.values() if s > 1)
    largest = sorted(
        ((cid, sz) for cid, sz in cluster_sizes.items() if sz > 1),
        key=lambda x: (-x[1], x[0]),
    )[:10]

    lines: list[tuple[str, str]] = [
        ("Input file", str(input_file.name)),
        ("", ""),
        ("START HERE", "Open 'Grouped Review' — each product once, matches nested below"),
        ("", ""),
        ("Total rows", f"{n:,}"),
        ("Similarity threshold", f"{similarity_threshold:.0%}"),
        ("Review Pairs min score", f"{review_pairs_min_score:.0%}"),
        ("Grouped Review rows", f"{len(grouped_review_df):,}"),
        ("Pairs >= threshold", f"{n_pairs:,}"),
        ("Exact pairs (score = 1.0)", f"{n_exact_pairs:,}"),
        ("Review Pairs sheet rows", f"{len(review_pairs_df):,}"),
        ("Exact duplicate groups", f"{n_dup_groups:,}"),
        ("Rows in exact duplicate groups", f"{rows_in_dups:,}"),
        ("Fuzzy clusters (size > 1)", f"{multi_clusters:,}"),
        ("Rows in fuzzy clusters", f"{rows_in_multi:,}"),
        ("Needs Review sheet rows", f"{len(needs_review_df):,}"),
        ("Singleton rows (no review needed)", f"{n - rows_in_multi:,}"),
        ("", ""),
        ("Largest clusters", "cluster_id | size"),
    ]
    for cid, sz in largest:
        lines.append((f"  Cluster {cid}", f"{sz:,} rows"))
    return pd.DataFrame(lines, columns=["metric", "value"])


def _style_header_row(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = TOP


def _apply_group_colors(ws, group_col_name: str, headers: list) -> None:
    if group_col_name not in headers:
        return
    gcol = headers.index(group_col_name) + 1
    ncols = len(headers)
    prev_group = object()
    palette_idx = -1
    for row in range(2, ws.max_row + 1):
        group_val = ws.cell(row=row, column=gcol).value
        if group_val != prev_group:
            palette_idx = (palette_idx + 1) % len(GROUP_FILLS)
            prev_group = group_val
        fill = GROUP_FILLS[palette_idx]
        for col in range(1, ncols + 1):
            ws.cell(row=row, column=col).fill = fill


def _set_column_widths(ws, widths: dict[str, float], headers: list) -> None:
    for name, width in widths.items():
        if name in headers:
            ws.column_dimensions[get_column_letter(headers.index(name) + 1)].width = width


def _hide_columns(ws, names: tuple[str, ...], headers: list) -> None:
    for name in names:
        if name in headers:
            ws.column_dimensions[get_column_letter(headers.index(name) + 1)].hidden = True


def _wrap_columns(ws, names: tuple[str, ...], headers: list) -> None:
    for name in names:
        if name not in headers:
            continue
        col_idx = headers.index(name) + 1
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_idx).alignment = WRAP


def format_workbook(wb) -> None:
    """Apply filters, freeze panes, column widths, and colour banding."""
    sheet_specs: dict[str, dict] = {
        "Summary": {
            "freeze": "A2",
            "widths": {"metric": 28, "value": 72},
            "wrap": ("value",),
        },
        "Grouped Review": {
            "freeze": "E2",
            "widths": {
                "cluster_id": 10, "cluster_size": 10, "position_in_cluster": 12,
                "depth": 8, "product_number": 16, "description": 52,
                "linked_to_product": 16, "score_to_parent": 12,
                "n_similar_in_cluster": 14, "exact_dup_group": 12,
            },
            "wrap": ("description",),
            "group_col": "cluster_id",
            "score_col": "score_to_parent",
        },
        "Review Pairs": {
            "freeze": "C2",
            "widths": {
                "score": 8, "exact": 8,
                "product_a": 16, "description_a": 44,
                "product_b": 16, "description_b": 44,
                "token_diff": 52,
            },
            "wrap": ("description_a", "description_b", "token_diff"),
            "hide": ("row_a", "row_b"),
            "score_col": "score",
            "exact_col": "exact",
        },
        "Exact Duplicates": {
            "freeze": "D2",
            "widths": {
                "dup_group_id": 12, "dup_group_size": 12,
                "product_number": 16, "description": 52,
            },
            "wrap": ("description",),
            "group_col": "dup_group_id",
        },
        "Needs Review": {
            "freeze": "D2",
            "widths": {
                "review_type": 16, "group_id": 10, "group_size": 10,
                "product_number": 16, "description": 52,
                "cluster_id": 10, "cluster_size": 10,
            },
            "wrap": ("description",),
            "group_col": "group_id",
        },
        "Clusters": {
            "freeze": "C2",
            "widths": {
                "product_number": 16, "description": 52,
                "cluster_id": 10, "cluster_size": 10,
            },
            "wrap": ("description",),
            "hide": ("row_index", "description_sorted"),
            "group_col": "cluster_id",
        },
        "Top Matches": {
            "freeze": "C2",
            "widths": {"product_number": 16, "description": 40},
            "wrap": ("description",),
            "hide": ("row_index", "description_sorted"),
        },
        "All Pairs": {
            "freeze": "C2",
            "widths": {
                "score": 8, "product_a": 16, "desc_a": 40,
                "product_b": 16, "desc_b": 40,
            },
            "wrap": ("desc_a", "desc_b"),
            "hide": ("row_a", "row_b", "desc_a_sorted", "desc_b_sorted"),
            "score_col": "score",
            "exact_col": "exact",
        },
        "All Pairs (top 1M)": {
            "freeze": "C2",
            "widths": {
                "score": 8, "product_a": 16, "desc_a": 40,
                "product_b": 16, "desc_b": 40,
            },
            "wrap": ("desc_a", "desc_b"),
            "hide": ("row_a", "row_b", "desc_a_sorted", "desc_b_sorted"),
            "score_col": "score",
            "exact_col": "exact",
        },
    }

    for ws in wb.worksheets:
        spec = sheet_specs.get(ws.title, {})
        if ws.max_row < 1 or ws.max_column < 1:
            continue
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        ncols = len(headers)

        _style_header_row(ws, ncols)
        ws.auto_filter.ref = ws.dimensions
        if freeze := spec.get("freeze"):
            ws.freeze_panes = freeze
        if widths := spec.get("widths"):
            _set_column_widths(ws, widths, headers)
        if wrap_cols := spec.get("wrap"):
            _wrap_columns(ws, wrap_cols, headers)
        if hide_cols := spec.get("hide"):
            _hide_columns(ws, hide_cols, headers)
        if group_col := spec.get("group_col"):
            _apply_group_colors(ws, group_col, headers)

        if score_col := spec.get("score_col"):
            if score_col in headers:
                letter = get_column_letter(headers.index(score_col) + 1)
                ws.conditional_formatting.add(
                    f"{letter}2:{letter}{ws.max_row}",
                    CellIsRule(operator="greaterThanOrEqual", formula=["0.9999"], fill=EXACT_FILL),
                )
                ws.conditional_formatting.add(
                    f"{letter}2:{letter}{ws.max_row}",
                    CellIsRule(operator="between", formula=["0.70", "0.9998"], fill=HIGH_FILL),
                )
        if exact_col := spec.get("exact_col"):
            if exact_col in headers:
                letter = get_column_letter(headers.index(exact_col) + 1)
                ws.conditional_formatting.add(
                    f"{letter}2:{letter}{ws.max_row}",
                    CellIsRule(operator="equal", formula=["True"], fill=EXACT_FILL),
                )


def resolve_input_file(cli_path: str | None = None) -> Path | None:
    """Pick the Excel file to read.

    Priority:
      1. Explicit CLI path (absolute or relative)
      2. CONFIG INPUT_FILE (as-is, then inside INPUT_FOLDER)
      3. Newest .xlsx inside INPUT_FOLDER (skip Excel lock files '~$...')
    """
    folder = Path(INPUT_FOLDER)
    folder.mkdir(parents=True, exist_ok=True)

    if cli_path:
        cand = Path(cli_path)
        if not cand.is_absolute():
            for probe in (cand, folder / cand, Path.cwd() / cand):
                if probe.exists():
                    return probe
        elif cand.exists():
            return cand
        print(f"ERROR: input file not found: {cli_path!r}")
        return None

    if INPUT_FILE:
        for cand in (Path(INPUT_FILE), folder / INPUT_FILE):
            if cand.exists():
                return cand
        print(f"ERROR: configured INPUT_FILE not found: {INPUT_FILE!r}")
        print(f"  Looked at: {Path(INPUT_FILE).resolve()}")
        print(f"  And     : {(folder / INPUT_FILE).resolve()}")
        return None

    candidates = [
        p for p in folder.glob("*.xlsx")
        if not p.name.startswith("~$")
    ]
    if not candidates:
        print(f"ERROR: no .xlsx file found in {folder.resolve()}")
        print("  Drop your Excel file in that folder and re-run,")
        print("  or: python similarity.py path\\to\\file.xlsx")
        return None
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    if len(candidates) > 1:
        print(f"  Found {len(candidates)} .xlsx files in {folder}; using the newest:")
        for p in candidates[:5]:
            print(f"    {p.name}")
    return candidates[0]


def parse_args(argv: list[str] | None = None) -> tuple[str | None, Path, bool]:
    """Return (optional input path, output path, gui_fast_mode).

    Usage:
        python similarity.py
        python similarity.py my_products.xlsx
        python similarity.py my_products.xlsx -o results_run2.xlsx
        python similarity.py FOExport.xlsx -o out.xlsx --gui
    """
    import argparse

    parser = argparse.ArgumentParser(
        description="Find near-duplicate product descriptions and write review Excel.",
    )
    parser.add_argument(
        "input",
        nargs="?",
        default=None,
        help="Input .xlsx path (or name inside the input/ folder). "
             "Default: newest file in input/, or INPUT_FILE in config.",
    )
    parser.add_argument(
        "-o", "--output",
        default=OUTPUT_FILE,
        help=f"Output .xlsx path (default: {OUTPUT_FILE})",
    )
    parser.add_argument(
        "--gui",
        action="store_true",
        help="Fast path for the review GUI: only write the sheets the reviewer needs.",
    )
    args = parser.parse_args(argv)
    return args.input, Path(args.output), bool(args.gui)


def main(argv: list[str] | None = None) -> int:
    t0 = time.time()
    try:
        sys.stdout.reconfigure(line_buffering=True)  # type: ignore[attr-defined]
    except Exception:
        pass
    cli_input, out_path, gui_mode = parse_args(argv)
    in_path = resolve_input_file(cli_input)
    if in_path is None:
        return 1

    print(f"Reading {in_path} ...")
    df = pd.read_excel(in_path, sheet_name=SHEET_NAME)

    try:
        prod_col, desc_col, search_col = resolve_columns(df)
    except ValueError as exc:
        print(f"ERROR: {exc}")
        return 1

    if detect_fo_export(df):
        print("  Format              : FOExport (Product number + Product name)")
    print(f"  Product column     : {prod_col!r}")
    print(f"  Description column : {desc_col!r}")
    if search_col:
        print(f"  Search-name fallback: {search_col!r}")
    print(f"  Rows               : {len(df):,}")

    products = [clean_product_id(v) for v in df[prod_col].tolist()]
    descriptions = build_descriptions(df, desc_col, search_col)

    tokenized = [tokenize(d) for d in descriptions]
    token_sets: list[frozenset[str]] = [t[0] for t in tokenized]
    canonical: list[str] = [t[1] for t in tokenized]
    n = len(token_sets)

    # ---- build inverted index ------------------------------------------------
    inv: dict[str, list[int]] = defaultdict(list)
    for doc_id, ts in enumerate(token_sets):
        for tok in ts:
            inv[tok].append(doc_id)

    if MAX_DOC_FREQUENCY > 0:
        cutoff = MAX_DOC_FREQUENCY * n
        dropped = {t for t, lst in inv.items() if len(lst) > cutoff}
        if dropped:
            for t in dropped:
                del inv[t]
            token_sets = [ts - dropped for ts in token_sets]
            # Re-derive canonical to reflect what's actually being compared.
            canonical = [" ".join(sorted(ts)) for ts in token_sets]
            print(f"  Dropped {len(dropped):,} ultra-common tokens (>{MAX_DOC_FREQUENCY:.0%} of rows)")

    hot = sorted(((len(v), k) for k, v in inv.items() if len(v) > 0.20 * n), reverse=True)[:5]
    if hot:
        print("  Note: these tokens still appear in >20% of rows:")
        for cnt, tok in hot:
            print(f"    {tok!r:>20s}  in {cnt:,} rows ({cnt / n:.0%})")
        if MAX_DOC_FREQUENCY <= 0:
            print("    Tip: set MAX_DOC_FREQUENCY = 0.5 at the top of this file to speed up.")

    print(f"  Vocabulary: {len(inv):,} unique tokens")
    # Full-row index was only needed for DF filtering / diagnostics.
    del inv

    # Collapse identical token sets so Jaccard is computed once per unique text,
    # then edges are expanded back to all member row ids.
    groups_by_set: dict[frozenset[str], list[int]] = defaultdict(list)
    for doc_id, ts in enumerate(token_sets):
        groups_by_set[ts].append(doc_id)

    unique_items: list[tuple[frozenset[str], list[int]]] = [
        (ts, ids) for ts, ids in groups_by_set.items() if ts
    ]
    u = len(unique_items)
    n_unique_incl_empty = len(groups_by_set)
    print(
        f"  Unique descriptions: {n_unique_incl_empty:,}  "
        f"(collapsed {n - n_unique_incl_empty:,} exact-text copies)"
    )
    print(f"Computing similarities (threshold={SIMILARITY_THRESHOLD}) ...")

    # Inverted index over unique descriptions only (postings stay sorted).
    inv_u: dict[str, list[int]] = defaultdict(list)
    for uid, (ts, _) in enumerate(unique_items):
        for tok in ts:
            inv_u[tok].append(uid)
    sizes_u = [len(ts) for ts, _ in unique_items]

    pair_edges: list[tuple[int, int, float]] = []   # only stores i < j

    # Exact score-1.0 edges within each identical-token-set group.
    for _, ids in unique_items:
        if len(ids) < 2:
            continue
        for a in range(len(ids)):
            ia = ids[a]
            for b in range(a + 1, len(ids)):
                pair_edges.append((ia, ids[b], 1.0))

    # Empty descriptions form one exact group (no tokens to index).
    empty_ids = groups_by_set.get(frozenset(), [])
    for a in range(len(empty_ids)):
        ia = empty_ids[a]
        for b in range(a + 1, len(empty_ids)):
            pair_edges.append((ia, empty_ids[b], 1.0))

    need_top = EMIT_WIDE_TOP_MATCHES
    top_matches: list[list[tuple[int, float]]] = [[] for _ in range(n)] if need_top else []
    # Score thresholds: Jaccard >= T  <=>  inter >= T/(1+T) * (size_i+size_j)
    # and size ratio must be at least T (smaller cannot be too small vs larger).
    thr = SIMILARITY_THRESHOLD
    thr_ratio = thr / (1.0 + thr) if thr < 1.0 else 0.5

    progress_step = max(1, u // 50)
    for uid_i in range(u):
        ts_i, ids_i = unique_items[uid_i]
        counts: Counter[int] = Counter()
        for tok in ts_i:
            postings = inv_u[tok]
            # Only candidates with higher unique-id → each pair scored once.
            start = bisect_right(postings, uid_i)
            if start < len(postings):
                counts.update(postings[start:])
        size_i = sizes_u[uid_i]
        for uid_j, inter in counts.items():
            size_j = sizes_u[uid_j]
            # Impossible to reach threshold if size ratio is too skewed.
            smaller, larger = (size_i, size_j) if size_i <= size_j else (size_j, size_i)
            if larger == 0 or smaller / larger < thr:
                continue
            if inter < thr_ratio * (size_i + size_j):
                continue
            union = size_i + size_j - inter
            jac = inter / union
            if jac < thr:
                continue
            ids_j = unique_items[uid_j][1]
            for i in ids_i:
                for j in ids_j:
                    a, b = (i, j) if i < j else (j, i)
                    pair_edges.append((a, b, jac))
                    if need_top:
                        top_matches[i].append((j, jac))
                        top_matches[j].append((i, jac))
        if (uid_i + 1) % progress_step == 0 or uid_i + 1 == u:
            elapsed = time.time() - t0
            rate = (uid_i + 1) / elapsed if elapsed else 0
            eta = (u - uid_i - 1) / rate if rate else 0
            print(f"  {uid_i + 1:>6,} / {u:,} unique   elapsed {elapsed:6.1f}s   eta {eta:6.1f}s")

    if need_top:
        # Exact copies also count as score-1.0 neighbours.
        for _, ids in unique_items:
            if len(ids) < 2:
                continue
            for a, i in enumerate(ids):
                for j in ids[a + 1 :]:
                    top_matches[i].append((j, 1.0))
                    top_matches[j].append((i, 1.0))
        for i in range(n):
            local = top_matches[i]
            local.sort(key=lambda x: -x[1])
            top_matches[i] = local[:TOP_N_MATCHES]

    # ---- union-find clustering ----------------------------------------------
    parent = list(range(n))

    def find(x: int) -> int:
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a: int, b: int) -> None:
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[ra] = rb

    for i, j, _ in pair_edges:
        union(i, j)

    roots = [find(i) for i in range(n)]
    root_sizes = Counter(roots)
    root_to_id = {r: new for new, r in enumerate(
        sorted(root_sizes, key=lambda r: (-root_sizes[r], r))
    )}
    cluster_ids = [root_to_id[r] for r in roots]
    cluster_sizes = {root_to_id[r]: s for r, s in root_sizes.items()}

    # ---- assemble output -----------------------------------------------------
    print("Assembling output ...")

    # Exact Duplicates — same canonical form = same bag of words.
    by_canonical: dict[str, list[int]] = defaultdict(list)
    for i, c in enumerate(canonical):
        if c:
            by_canonical[c].append(i)
    dup_groups = sorted(
        ((c, ids) for c, ids in by_canonical.items() if len(ids) > 1),
        key=lambda x: (-len(x[1]), x[0]),
    )
    dup_rows = []
    for gid, (_canon, ids) in enumerate(dup_groups):
        for i in ids:
            dup_rows.append({
                "dup_group_id": gid,
                "dup_group_size": len(ids),
                "product_number": products[i],
                "description": descriptions[i],
            })
    dup_df = pd.DataFrame(dup_rows) if dup_rows else pd.DataFrame(
        columns=["dup_group_id", "dup_group_size", "product_number", "description"]
    )

    dup_group_by_row = {
        i: gid for gid, (_canon, ids) in enumerate(dup_groups) for i in ids
    }

    grouped_review_df = build_grouped_review_df(
        products,
        descriptions,
        cluster_ids,
        cluster_sizes,
        pair_edges,
        dup_group_by_row,
    )

    semantic_df: pd.DataFrame | None = None
    if SEMANTIC_ENABLED:
        try:
            semantic_df = build_semantic_suggestions_df(
                products, descriptions, cluster_ids,
            )
        except Exception as exc:  # noqa: BLE001 - never fail the run over hints
            print(f"  Semantic pass failed: {exc}")
            semantic_df = None

    review_pairs_df = None
    needs_review_df = None
    cluster_df = None
    if not gui_mode:
        review_pairs_df = build_review_pairs_df(
            pair_edges, products, descriptions, token_sets, REVIEW_PAIRS_MIN_SCORE,
        )
        needs_review_df = build_needs_review_df(
            products, descriptions, cluster_ids, cluster_sizes, dup_groups,
        )
        cluster_df = (
            pd.DataFrame({
                "row_index": list(range(n)),
                "product_number": products,
                "description": descriptions,
                "description_sorted": canonical,
                "cluster_id": cluster_ids,
                "cluster_size": [cluster_sizes[c] for c in cluster_ids],
            })
            .sort_values(
                ["cluster_size", "cluster_id", "description", "row_index"],
                ascending=[False, True, True, True],
            )
            .reset_index(drop=True)
        )
    else:
        # Lightweight placeholders for summary metrics in GUI mode.
        review_pairs_df = pd.DataFrame()
        needs_review_df = grouped_review_df
        print("  GUI mode: skipping Review Pairs / Needs Review / Clusters sheets")

    summary_df = build_summary_df(
        n=n,
        pair_edges=pair_edges,
        review_pairs_df=review_pairs_df,
        grouped_review_df=grouped_review_df,
        dup_groups=dup_groups,
        cluster_sizes=cluster_sizes,
        needs_review_df=needs_review_df,
        similarity_threshold=SIMILARITY_THRESHOLD,
        review_pairs_min_score=REVIEW_PAIRS_MIN_SCORE,
        input_file=in_path,
    )

    # Legacy wide Top Matches sheet (optional).
    top_df: pd.DataFrame | None = None
    if EMIT_WIDE_TOP_MATCHES:
        out_rows = []
        for i, matches in enumerate(top_matches):
            row = {
                "row_index": i,
                "product_number": products[i],
                "description": descriptions[i],
                "description_sorted": canonical[i],
                "cluster_id": cluster_ids[i],
                "cluster_size": cluster_sizes[cluster_ids[i]],
                "n_matches": len(matches),
            }
            for k, (j, jac) in enumerate(matches, start=1):
                row[f"match{k}_row"] = j
                row[f"match{k}_product"] = products[j]
                row[f"match{k}_score"] = round(jac, 4)
                row[f"match{k}_exact"] = jac >= 0.9999
                row[f"match{k}_desc"] = descriptions[j]
                row[f"match{k}_desc_sorted"] = canonical[j]
            out_rows.append(row)
        top_df = pd.DataFrame(out_rows)

    # Legacy All Pairs sheet (optional).
    pair_df: pd.DataFrame | None = None
    truncated = False
    if EMIT_ALL_PAIRS_SHEET:
        EXCEL_ROW_CAP = 1_000_000
        pair_df = pd.DataFrame(pair_edges, columns=["row_a", "row_b", "score"])
        if not pair_df.empty:
            pair_df = pair_df.sort_values("score", ascending=False).reset_index(drop=True)
            pair_df["score"] = pair_df["score"].round(4)
            pair_df["product_a"] = pair_df["row_a"].map(lambda i: products[i])
            pair_df["desc_a"] = pair_df["row_a"].map(lambda i: descriptions[i])
            pair_df["desc_a_sorted"] = pair_df["row_a"].map(lambda i: canonical[i])
            pair_df["product_b"] = pair_df["row_b"].map(lambda i: products[i])
            pair_df["desc_b"] = pair_df["row_b"].map(lambda i: descriptions[i])
            pair_df["desc_b_sorted"] = pair_df["row_b"].map(lambda i: canonical[i])
            pair_df["exact"] = pair_df["score"] >= 0.9999
            pair_df = pair_df[[
                "score", "exact",
                "row_a", "product_a", "desc_a", "desc_a_sorted",
                "row_b", "product_b", "desc_b", "desc_b_sorted",
            ]]
            if len(pair_df) > EXCEL_ROW_CAP:
                truncated = True
                full_pair_csv = out_path.with_suffix(".all_pairs.csv")
                pair_df.to_csv(full_pair_csv, index=False)
                print(f"  All Pairs has {len(pair_df):,} rows; full list written to {full_pair_csv}")
                pair_df = pair_df.head(EXCEL_ROW_CAP)

    print(f"Writing {out_path} ...")
    with pd.ExcelWriter(out_path, engine="openpyxl") as xl:
        summary_df.to_excel(xl, sheet_name="Summary", index=False)
        grouped_review_df.to_excel(xl, sheet_name="Grouped Review", index=False)
        dup_df.to_excel(xl, sheet_name="Exact Duplicates", index=False)
        if semantic_df is not None:
            semantic_df.to_excel(xl, sheet_name="Semantic Suggestions", index=False)
        if not gui_mode:
            review_pairs_df.to_excel(xl, sheet_name="Review Pairs", index=False)
            needs_review_df.to_excel(xl, sheet_name="Needs Review", index=False)
            cluster_df.to_excel(xl, sheet_name="Clusters", index=False)
            if top_df is not None:
                top_df.to_excel(xl, sheet_name="Top Matches", index=False)
            if pair_df is not None:
                sheet = "All Pairs (top 1M)" if truncated else "All Pairs"
                pair_df.to_excel(xl, sheet_name=sheet, index=False)
            format_workbook(xl.book)
        else:
            # Keep GUI runs fast: skip per-cell wrapping/colours on huge sheets.
            for ws in xl.book.worksheets:
                if ws.max_row < 1 or ws.max_column < 1:
                    continue
                _style_header_row(ws, ws.max_column)
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions

    multi_clusters = sum(1 for s in cluster_sizes.values() if s > 1)
    rows_in_multi = sum(s for s in cluster_sizes.values() if s > 1)
    n_dup_groups = len(dup_groups)
    rows_in_dups = sum(len(ids) for _, ids in dup_groups)
    print()
    print(f"Done in {time.time() - t0:.1f}s")
    print(f"  Rows                          : {n:,}")
    print(f"  Pairs >= threshold            : {len(pair_edges):,}")
    print(f"  Grouped Review rows           : {len(grouped_review_df):,}")
    if not gui_mode:
        print(f"  Review Pairs rows             : {len(review_pairs_df):,}")
        print(f"  Needs Review rows             : {len(needs_review_df):,}")
    print(f"  Fuzzy clusters (size > 1)     : {multi_clusters:,}")
    print(f"  Rows in such clusters         : {rows_in_multi:,}")
    print(f"  Exact-duplicate groups        : {n_dup_groups:,}")
    print(f"  Rows in exact-duplicate groups: {rows_in_dups:,}")
    if semantic_df is not None:
        print(f"  Semantic suggestion rows      : {len(semantic_df):,}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
